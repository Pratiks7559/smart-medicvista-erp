from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Q
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from datetime import datetime, timedelta
import csv
import json
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import ProductMaster
from .stock_manager import StockManager


@login_required
def stock_statement_report(request):
    """
    Comprehensive stock statement report showing opening, received, sold, and balance
    """
    # Get filter parameters
    search_query = request.GET.get('search', '').strip()
    category_filter = request.GET.get('category', '')
    company_filter = request.GET.get('company', '')
    stock_status = request.GET.get('stock_status', '')  # all, in_stock, low_stock, out_of_stock
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Base query
    products_query = ProductMaster.objects.all().order_by('product_name')
    
    # Apply filters
    if search_query:
        products_query = products_query.filter(
            Q(product_name__icontains=search_query) |
            Q(product_company__icontains=search_query) |
            Q(product_salt__icontains=search_query) |
            Q(product_barcode__icontains=search_query)
        )
    
    if category_filter:
        products_query = products_query.filter(product_category__icontains=category_filter)
    
    if company_filter:
        products_query = products_query.filter(product_company__icontains=company_filter)
    
    # Get unique categories and companies for filter dropdowns
    categories = ProductMaster.objects.values_list('product_category', flat=True).distinct().order_by('product_category')
    companies = ProductMaster.objects.values_list('product_company', flat=True).distinct().order_by('product_company')
    
    # Pagination
    paginator = Paginator(products_query, 25)  # 25 products per page
    page_number = request.GET.get('page')
    products_page = paginator.get_page(page_number)
    
    # Process stock data for each product
    stock_data = []
    total_opening = 0
    total_received = 0
    total_sold = 0
    total_balance = 0
    total_value = 0
    
    for product in products_page:
        try:
            # Get comprehensive stock summary
            stock_summary = StockManager.get_stock_summary(product.productid)
            
            # Calculate opening stock (this would need historical data - for now using current stock)
            # In a real scenario, you'd calculate this based on a specific date
            opening_stock = 0  # This should be calculated based on date_from parameter
            
            received_stock = stock_summary['total_purchased'] + stock_summary['total_sales_returns']
            sold_stock = stock_summary['total_sold'] + stock_summary['total_purchase_returns']
            balance_stock = stock_summary['total_stock']
            
            # Get average MRP for value calculation
            avg_mrp = 0
            if stock_summary['batches']:
                mrp_values = []
                for batch in stock_summary['batches']:
                    try:
                        # Get MRP from purchase records for this batch
                        from .models import PurchaseMaster
                        purchase = PurchaseMaster.objects.filter(
                            productid=product.productid,
                            product_batch_no=batch['batch_no']
                        ).first()
                        if purchase and purchase.product_MRP:
                            mrp_values.append(float(purchase.product_MRP))
                    except:
                        continue
                
                if mrp_values:
                    avg_mrp = sum(mrp_values) / len(mrp_values)
            
            stock_value = balance_stock * avg_mrp
            
            # Determine stock status
            if balance_stock <= 0:
                status = 'out_of_stock'
                status_label = 'Out of Stock'
                status_class = 'danger'
            elif balance_stock < 10:
                status = 'low_stock'
                status_label = 'Low Stock'
                status_class = 'warning'
            else:
                status = 'in_stock'
                status_label = 'In Stock'
                status_class = 'success'
            
            # Apply stock status filter
            if stock_status and stock_status != 'all' and stock_status != status:
                continue
            
            stock_item = {
                'product': product,
                'opening_stock': opening_stock,
                'received_stock': received_stock,
                'sold_stock': sold_stock,
                'balance_stock': balance_stock,
                'avg_mrp': avg_mrp,
                'stock_value': stock_value,
                'status': status,
                'status_label': status_label,
                'status_class': status_class,
                'batches': stock_summary['batches']
            }
            
            stock_data.append(stock_item)
            
            # Add to totals
            total_opening += opening_stock
            total_received += received_stock
            total_sold += sold_stock
            total_balance += balance_stock
            total_value += stock_value
            
        except Exception as e:
            print(f"Error processing stock for {product.product_name}: {e}")
            # Add error entry
            stock_data.append({
                'product': product,
                'opening_stock': 0,
                'received_stock': 0,
                'sold_stock': 0,
                'balance_stock': 0,
                'avg_mrp': 0,
                'stock_value': 0,
                'status': 'error',
                'status_label': 'Error',
                'status_class': 'secondary',
                'batches': []
            })
    
    # Handle AJAX requests for export
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        export_type = request.GET.get('export')
        
        if export_type == 'pdf':
            return export_stock_statement_pdf(request, stock_data)
        elif export_type == 'excel':
            return export_stock_statement_excel(request, stock_data)
    
    context = {
        'stock_data': stock_data,
        'products_page': products_page,
        'search_query': search_query,
        'category_filter': category_filter,
        'company_filter': company_filter,
        'stock_status': stock_status,
        'date_from': date_from,
        'date_to': date_to,
        'categories': [cat for cat in categories if cat],
        'companies': [comp for comp in companies if comp],
        'totals': {
            'opening': total_opening,
            'received': total_received,
            'sold': total_sold,
            'balance': total_balance,
            'value': total_value
        },
        'title': 'Stock Statement Report'
    }
    
    return render(request, 'reports/stock_statement_report.html', context)


@login_required
def export_stock_statement_pdf(request):
    """Export stock statement to PDF"""
    # Get all products
    products_query = ProductMaster.objects.all().order_by('product_name')
    
    # Process stock data
    stock_data = []
    totals = {'opening': 0, 'received': 0, 'sold': 0, 'balance': 0, 'value': 0}
    
    for product in products_query:
        try:
            stock_summary = StockManager.get_stock_summary(product.productid)
            opening_stock = 0
            received_stock = stock_summary['total_purchased'] + stock_summary['total_sales_returns']
            sold_stock = stock_summary['total_sold'] + stock_summary['total_purchase_returns']
            balance_stock = stock_summary['total_stock']
            
            avg_mrp = 0
            if stock_summary['batches']:
                from .models import PurchaseMaster
                mrp_values = []
                for batch in stock_summary['batches']:
                    purchase = PurchaseMaster.objects.filter(
                        productid=product.productid,
                        product_batch_no=batch['batch_no']
                    ).first()
                    if purchase and purchase.product_MRP:
                        mrp_values.append(float(purchase.product_MRP))
                if mrp_values:
                    avg_mrp = sum(mrp_values) / len(mrp_values)
            
            stock_value = balance_stock * avg_mrp
            
            if balance_stock <= 0:
                status_label = 'Out of Stock'
            elif balance_stock < 10:
                status_label = 'Low Stock'
            else:
                status_label = 'In Stock'
            
            stock_data.append({
                'product': product,
                'opening_stock': opening_stock,
                'received_stock': received_stock,
                'sold_stock': sold_stock,
                'balance_stock': balance_stock,
                'avg_mrp': avg_mrp,
                'stock_value': stock_value,
                'status_label': status_label
            })
            
            totals['opening'] += opening_stock
            totals['received'] += received_stock
            totals['sold'] += sold_stock
            totals['balance'] += balance_stock
            totals['value'] += stock_value
        except:
            continue
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="stock_statement_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    elements = []
    styles = getSampleStyleSheet()
    
    title = Paragraph("<b>Stock Statement Report</b>", ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=1))
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    date_text = f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    elements.append(Paragraph(date_text, ParagraphStyle('Date', parent=styles['Normal'], fontSize=9, alignment=1)))
    elements.append(Spacer(1, 15))
    
    # Summary
    summary_data = [[
        f"Opening: {int(totals['opening'])}",
        f"Received: {int(totals['received'])}",
        f"Sold: {int(totals['sold'])}",
        f"Balance: {int(totals['balance'])}",
        f"Value: ₹{totals['value']:.2f}"
    ]]
    summary_table = Table(summary_data, colWidths=[1.8*inch]*5)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 15))
    
    # Main table
    data = [['Product', 'Company', 'Opening', 'Received', 'Sold', 'Balance', 'MRP', 'Value', 'Status']]
    for item in stock_data:
        data.append([
            item['product'].product_name[:30],
            item['product'].product_company[:15],
            str(int(item['opening_stock'])),
            str(int(item['received_stock'])),
            str(int(item['sold_stock'])),
            str(int(item['balance_stock'])),
            f"₹{item['avg_mrp']:.2f}",
            f"₹{item['stock_value']:.2f}",
            item['status_label']
        ])
    
    table = Table(data, colWidths=[2.5*inch, 1.3*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.8*inch, 1*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    elements.append(table)
    
    doc.build(elements)
    return response


def export_stock_statement_excel(request, stock_data):
    """Export stock statement to Excel"""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'stock_statement_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Stock Statement'
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Add headers
    headers = [
        'Product Name', 'Company', 'Category', 'Packing', 'Opening Stock',
        'Received', 'Sold', 'Balance', 'Avg MRP', 'Stock Value', 'Status'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data
    for row, item in enumerate(stock_data, 2):
        worksheet.cell(row=row, column=1, value=item['product'].product_name)
        worksheet.cell(row=row, column=2, value=item['product'].product_company)
        worksheet.cell(row=row, column=3, value=item['product'].product_category)
        worksheet.cell(row=row, column=4, value=item['product'].product_packing)
        worksheet.cell(row=row, column=5, value=item['opening_stock'])
        worksheet.cell(row=row, column=6, value=item['received_stock'])
        worksheet.cell(row=row, column=7, value=item['sold_stock'])
        worksheet.cell(row=row, column=8, value=item['balance_stock'])
        worksheet.cell(row=row, column=9, value=item['avg_mrp'])
        worksheet.cell(row=row, column=10, value=item['stock_value'])
        worksheet.cell(row=row, column=11, value=item['status_label'])
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(response)
    return response


@login_required
def stock_statement_batch_detail(request, product_id):
    """Get detailed batch information for a product"""
    try:
        product = ProductMaster.objects.get(productid=product_id)
        stock_summary = StockManager.get_stock_summary(product_id)
        
        batch_details = []
        for batch in stock_summary['batches']:
            # Get additional batch information
            from .models import PurchaseMaster, SaleRateMaster
            
            # Get MRP and rates
            purchase = PurchaseMaster.objects.filter(
                productid=product_id,
                product_batch_no=batch['batch_no']
            ).first()
            
            try:
                rates = SaleRateMaster.objects.get(
                    productid=product_id,
                    product_batch_no=batch['batch_no']
                )
                rate_A = rates.rate_A
                rate_B = rates.rate_B
                rate_C = rates.rate_C
            except SaleRateMaster.DoesNotExist:
                rate_A = rate_B = rate_C = 0
            
            batch_details.append({
                'batch_no': batch['batch_no'],
                'expiry': batch['expiry'],
                'stock': batch['stock'],
                'purchased': batch['purchased'],
                'sold': batch['sold'],
                'purchase_returns': batch['purchase_returns'],
                'sales_returns': batch['sales_returns'],
                'mrp': purchase.product_MRP if purchase else 0,
                'rate_A': rate_A,
                'rate_B': rate_B,
                'rate_C': rate_C
            })
        
        return JsonResponse({
            'success': True,
            'product_name': product.product_name,
            'product_company': product.product_company,
            'batches': batch_details
        })
        
    except ProductMaster.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Product not found'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })