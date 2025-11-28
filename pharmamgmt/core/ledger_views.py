from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.utils import timezone
from .models import (
    CustomerMaster, SupplierMaster, SalesInvoiceMaster, SalesMaster,
    SalesInvoicePaid, ReturnSalesInvoiceMaster, InvoiceMaster, InvoicePaid,
    ReturnInvoiceMaster, Pharmacy_Details
)

@login_required
def ledger_selection(request):
    """View for ledger selection page"""
    context = {
        'title': 'Ledger Selection'
    }
    return render(request, 'ledger/ledger_selection.html', context)

@login_required
def customer_ledger(request, customer_id=None):
    customers = CustomerMaster.objects.all().order_by('customer_name')
    
    if not customer_id:
        context = {'customers': customers, 'title': 'Select Customer'}
        return render(request, 'ledger/customer_select.html', context)
    
    customer = get_object_or_404(CustomerMaster, customerid=customer_id)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    transactions = []
    
    # Sales Invoices (Debit)
    sales = SalesInvoiceMaster.objects.filter(customerid=customer).order_by('sales_invoice_date')
    if start_date and end_date:
        sales = sales.filter(sales_invoice_date__range=[start_date, end_date])
    
    for sale in sales:
        total = SalesMaster.objects.filter(sales_invoice_no=sale.sales_invoice_no).aggregate(
            Sum('sale_total_amount'))['sale_total_amount__sum'] or 0
        transactions.append({
            'date': sale.sales_invoice_date,
            'type': 'Sales Invoice',
            'reference': sale.sales_invoice_no,
            'debit': total,
            'credit': 0,
            'invoice_obj': sale
        })
    
    # Payments (Credit)
    payments = SalesInvoicePaid.objects.filter(
        sales_ip_invoice_no__customerid=customer
    ).order_by('sales_payment_date')
    if start_date and end_date:
        payments = payments.filter(sales_payment_date__range=[start_date, end_date])
    
    for payment in payments:
        transactions.append({
            'date': payment.sales_payment_date,
            'type': 'Payment',
            'reference': payment.sales_ip_invoice_no.sales_invoice_no,
            'debit': 0,
            'credit': payment.sales_payment_amount,
            'invoice_obj': payment.sales_ip_invoice_no
        })
    
    # Sales Returns (Credit)
    returns = ReturnSalesInvoiceMaster.objects.filter(
        return_sales_customerid=customer
    ).order_by('return_sales_invoice_date')
    if start_date and end_date:
        returns = returns.filter(return_sales_invoice_date__range=[start_date, end_date])
    
    for ret in returns:
        transactions.append({
            'date': ret.return_sales_invoice_date,
            'type': 'Sales Return',
            'reference': ret.return_sales_invoice_no,
            'debit': 0,
            'credit': ret.return_sales_invoice_total,
            'return_obj': ret
        })
    
    # Sort and calculate balance
    transactions.sort(key=lambda x: x['date'])
    balance = 0
    for trans in transactions:
        balance += trans['debit'] - trans['credit']
        trans['balance'] = balance
    
    total_debit = sum(t['debit'] for t in transactions)
    total_credit = sum(t['credit'] for t in transactions)
    
    # Get pharmacy details for print format
    try:
        pharmacy = Pharmacy_Details.objects.first()
    except Pharmacy_Details.DoesNotExist:
        pharmacy = None
    
    context = {
        'customer': customer,
        'transactions': transactions,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'closing_balance': balance,
        'customers': customers,
        'pharmacy': pharmacy,
        'current_date': timezone.now(),
        'start_date': start_date,
        'end_date': end_date,
        'title': f'Ledger - {customer.customer_name}'
    }
    return render(request, 'ledger/customer_ledger.html', context)


@login_required
def supplier_ledger(request, supplier_id=None):
    suppliers = SupplierMaster.objects.all().order_by('supplier_name')
    
    if not supplier_id:
        context = {'suppliers': suppliers, 'title': 'Select Supplier'}
        return render(request, 'ledger/supplier_select.html', context)
    
    supplier = get_object_or_404(SupplierMaster, supplierid=supplier_id)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    transactions = []
    
    # Purchase Invoices (Credit)
    purchases = InvoiceMaster.objects.filter(supplierid=supplier).order_by('invoice_date')
    if start_date and end_date:
        purchases = purchases.filter(invoice_date__range=[start_date, end_date])
    
    for purchase in purchases:
        transactions.append({
            'date': purchase.invoice_date,
            'type': 'Purchase Invoice',
            'reference': purchase.invoice_no,
            'debit': 0,
            'credit': purchase.invoice_total,
            'invoice_obj': purchase
        })
    
    # Payments (Debit)
    payments = InvoicePaid.objects.filter(
        ip_invoiceid__supplierid=supplier
    ).order_by('payment_date')
    if start_date and end_date:
        payments = payments.filter(payment_date__range=[start_date, end_date])
    
    for payment in payments:
        transactions.append({
            'date': payment.payment_date,
            'type': 'Payment',
            'reference': payment.ip_invoiceid.invoice_no,
            'debit': payment.payment_amount,
            'credit': 0,
            'invoice_obj': payment.ip_invoiceid
        })
    
    # Purchase Returns (Debit)
    returns = ReturnInvoiceMaster.objects.filter(
        returnsupplierid=supplier
    ).order_by('returninvoice_date')
    if start_date and end_date:
        returns = returns.filter(returninvoice_date__range=[start_date, end_date])
    
    for ret in returns:
        transactions.append({
            'date': ret.returninvoice_date,
            'type': 'Purchase Return',
            'reference': ret.returninvoiceid,
            'debit': ret.returninvoice_total,
            'credit': 0,
            'return_obj': ret
        })
    
    # Sort and calculate balance
    transactions.sort(key=lambda x: x['date'])
    balance = 0
    for trans in transactions:
        balance += trans['credit'] - trans['debit']
        trans['balance'] = balance
    
    total_debit = sum(t['debit'] for t in transactions)
    total_credit = sum(t['credit'] for t in transactions)
    
    # Get pharmacy details for print format
    try:
        pharmacy = Pharmacy_Details.objects.first()
    except Pharmacy_Details.DoesNotExist:
        pharmacy = None
    
    context = {
        'supplier': supplier,
        'transactions': transactions,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'closing_balance': balance,
        'suppliers': suppliers,
        'pharmacy': pharmacy,
        'current_date': timezone.now(),
        'start_date': start_date,
        'end_date': end_date,
        'title': f'Ledger - {supplier.supplier_name}'
    }
    return render(request, 'ledger/supplier_ledger.html', context)

@login_required
def export_supplier_ledger_pdf(request, supplier_id):
    """Export supplier ledger as PDF"""
    from django.http import HttpResponse
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from io import BytesIO
    
    supplier = get_object_or_404(SupplierMaster, supplierid=supplier_id)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Get transactions (same logic as supplier_ledger view)
    transactions = []
    
    # Purchase Invoices
    purchases = InvoiceMaster.objects.filter(supplierid=supplier).order_by('invoice_date')
    if start_date and end_date:
        purchases = purchases.filter(invoice_date__range=[start_date, end_date])
    
    for purchase in purchases:
        transactions.append({
            'date': purchase.invoice_date,
            'type': 'Purchase Invoice',
            'reference': purchase.invoice_no,
            'debit': 0,
            'credit': purchase.invoice_total
        })
    
    # Payments
    payments = InvoicePaid.objects.filter(ip_invoiceid__supplierid=supplier).order_by('payment_date')
    if start_date and end_date:
        payments = payments.filter(payment_date__range=[start_date, end_date])
    
    for payment in payments:
        transactions.append({
            'date': payment.payment_date,
            'type': 'Payment',
            'reference': payment.ip_invoiceid.invoice_no,
            'debit': payment.payment_amount,
            'credit': 0
        })
    
    # Sort and calculate balance
    transactions.sort(key=lambda x: x['date'])
    balance = 0
    for trans in transactions:
        balance += trans['credit'] - trans['debit']
        trans['balance'] = balance
    
    total_debit = sum(t['debit'] for t in transactions)
    total_credit = sum(t['credit'] for t in transactions)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title = Paragraph(f"Supplier Ledger - {supplier.supplier_name}", styles['Title'])
    story.append(title)
    story.append(Spacer(1, 12))
    
    # Supplier info
    supplier_info = Paragraph(f"<b>{supplier.supplier_name}</b><br/>Mobile: {supplier.supplier_mobile}", styles['Normal'])
    story.append(supplier_info)
    story.append(Spacer(1, 12))
    
    # Table data
    data = [['Date', 'Type', 'Invoice No', 'Debit (Paid)', 'Credit (Purchase)', 'Balance']]
    
    for trans in transactions:
        data.append([
            trans['date'].strftime('%d-%m-%Y'),
            trans['type'],
            trans['reference'],
            f"₹{trans['debit']:.2f}" if trans['debit'] > 0 else '-',
            f"₹{trans['credit']:.2f}" if trans['credit'] > 0 else '-',
            f"₹{trans['balance']:.2f}"
        ])
    
    # Total row
    data.append(['TOTAL', '', '', f"₹{total_debit:.2f}", f"₹{total_credit:.2f}", f"₹{balance:.2f}"])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.beige),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    doc.build(story)
    
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    
    # Check if download parameter is present
    if request.GET.get('download'):
        response['Content-Disposition'] = f'attachment; filename="supplier_ledger_{supplier.supplier_name}.pdf"'
    else:
        response['Content-Disposition'] = f'inline; filename="supplier_ledger_{supplier.supplier_name}.pdf"'
    
    return response

@login_required
def export_supplier_ledger_excel(request, supplier_id):
    """Export supplier ledger as Excel"""
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    
    supplier = get_object_or_404(SupplierMaster, supplierid=supplier_id)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Get transactions (same logic as supplier_ledger view)
    transactions = []
    
    # Purchase Invoices
    purchases = InvoiceMaster.objects.filter(supplierid=supplier).order_by('invoice_date')
    if start_date and end_date:
        purchases = purchases.filter(invoice_date__range=[start_date, end_date])
    
    for purchase in purchases:
        transactions.append({
            'date': purchase.invoice_date,
            'type': 'Purchase Invoice',
            'reference': purchase.invoice_no,
            'debit': 0,
            'credit': purchase.invoice_total
        })
    
    # Payments
    payments = InvoicePaid.objects.filter(ip_invoiceid__supplierid=supplier).order_by('payment_date')
    if start_date and end_date:
        payments = payments.filter(payment_date__range=[start_date, end_date])
    
    for payment in payments:
        transactions.append({
            'date': payment.payment_date,
            'type': 'Payment',
            'reference': payment.ip_invoiceid.invoice_no,
            'debit': payment.payment_amount,
            'credit': 0
        })
    
    # Sort and calculate balance
    transactions.sort(key=lambda x: x['date'])
    balance = 0
    for trans in transactions:
        balance += trans['credit'] - trans['debit']
        trans['balance'] = balance
    
    total_debit = sum(t['debit'] for t in transactions)
    total_credit = sum(t['credit'] for t in transactions)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Supplier Ledger"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center")
    
    # Title
    ws['A1'] = f"Supplier Ledger - {supplier.supplier_name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    # Supplier info
    ws['A2'] = f"Mobile: {supplier.supplier_mobile}"
    ws.merge_cells('A2:F2')
    
    # Headers
    headers = ['Date', 'Type', 'Invoice No', 'Debit (Paid)', 'Credit (Purchase)', 'Balance']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Data
    for row, trans in enumerate(transactions, 5):
        ws.cell(row=row, column=1, value=trans['date'].strftime('%d-%m-%Y'))
        ws.cell(row=row, column=2, value=trans['type'])
        ws.cell(row=row, column=3, value=trans['reference'])
        ws.cell(row=row, column=4, value=trans['debit'] if trans['debit'] > 0 else 0)
        ws.cell(row=row, column=5, value=trans['credit'] if trans['credit'] > 0 else 0)
        ws.cell(row=row, column=6, value=trans['balance'])
    
    # Total row
    total_row = len(transactions) + 5
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=total_debit).font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=total_credit).font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=balance).font = Font(bold=True)
    
    # Auto-adjust column widths
    for col_num in range(1, 7):  # 6 columns (A to F)
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)
        for row in ws.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                if hasattr(cell, 'value') and cell.value is not None:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="supplier_ledger_{supplier.supplier_name}.xlsx"'
    return response

@login_required
def export_customer_ledger_pdf(request, customer_id):
    """Export customer ledger as PDF"""
    from django.http import HttpResponse
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from io import BytesIO
    
    customer = get_object_or_404(CustomerMaster, customerid=customer_id)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    transactions = []
    
    # Sales Invoices
    sales = SalesInvoiceMaster.objects.filter(customerid=customer).order_by('sales_invoice_date')
    if start_date and end_date:
        sales = sales.filter(sales_invoice_date__range=[start_date, end_date])
    
    for sale in sales:
        total = SalesMaster.objects.filter(sales_invoice_no=sale.sales_invoice_no).aggregate(
            Sum('sale_total_amount'))['sale_total_amount__sum'] or 0
        transactions.append({
            'date': sale.sales_invoice_date,
            'type': 'Sales Invoice',
            'reference': sale.sales_invoice_no,
            'debit': total,
            'credit': 0
        })
    
    # Payments
    payments = SalesInvoicePaid.objects.filter(sales_ip_invoice_no__customerid=customer).order_by('sales_payment_date')
    if start_date and end_date:
        payments = payments.filter(sales_payment_date__range=[start_date, end_date])
    
    for payment in payments:
        transactions.append({
            'date': payment.sales_payment_date,
            'type': 'Payment',
            'reference': payment.sales_ip_invoice_no.sales_invoice_no,
            'debit': 0,
            'credit': payment.sales_payment_amount
        })
    
    # Sort and calculate balance
    transactions.sort(key=lambda x: x['date'])
    balance = 0
    for trans in transactions:
        balance += trans['debit'] - trans['credit']
        trans['balance'] = balance
    
    total_debit = sum(t['debit'] for t in transactions)
    total_credit = sum(t['credit'] for t in transactions)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title = Paragraph(f"Customer Ledger - {customer.customer_name}", styles['Title'])
    story.append(title)
    story.append(Spacer(1, 12))
    
    # Customer info
    customer_info = Paragraph(f"<b>{customer.customer_name}</b><br/>Mobile: {customer.customer_mobile}", styles['Normal'])
    story.append(customer_info)
    story.append(Spacer(1, 12))
    
    # Table data
    data = [['Date', 'Type', 'Invoice No', 'Debit (Sales)', 'Credit (Received)', 'Balance']]
    
    for trans in transactions:
        data.append([
            trans['date'].strftime('%d-%m-%Y'),
            trans['type'],
            trans['reference'],
            f"₹{trans['debit']:.2f}" if trans['debit'] > 0 else '-',
            f"₹{trans['credit']:.2f}" if trans['credit'] > 0 else '-',
            f"₹{trans['balance']:.2f}"
        ])
    
    # Total row
    data.append(['TOTAL', '', '', f"₹{total_debit:.2f}", f"₹{total_credit:.2f}", f"₹{balance:.2f}"])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.beige),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    doc.build(story)
    
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    
    # Check if download parameter is present
    if request.GET.get('download'):
        response['Content-Disposition'] = f'attachment; filename="customer_ledger_{customer.customer_name}.pdf"'
    else:
        response['Content-Disposition'] = f'inline; filename="customer_ledger_{customer.customer_name}.pdf"'
    
    return response

@login_required
def export_customer_ledger_excel(request, customer_id):
    """Export customer ledger as Excel"""
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    
    customer = get_object_or_404(CustomerMaster, customerid=customer_id)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    transactions = []
    
    # Sales Invoices
    sales = SalesInvoiceMaster.objects.filter(customerid=customer).order_by('sales_invoice_date')
    if start_date and end_date:
        sales = sales.filter(sales_invoice_date__range=[start_date, end_date])
    
    for sale in sales:
        total = SalesMaster.objects.filter(sales_invoice_no=sale.sales_invoice_no).aggregate(
            Sum('sale_total_amount'))['sale_total_amount__sum'] or 0
        transactions.append({
            'date': sale.sales_invoice_date,
            'type': 'Sales Invoice',
            'reference': sale.sales_invoice_no,
            'debit': total,
            'credit': 0
        })
    
    # Payments
    payments = SalesInvoicePaid.objects.filter(sales_ip_invoice_no__customerid=customer).order_by('sales_payment_date')
    if start_date and end_date:
        payments = payments.filter(sales_payment_date__range=[start_date, end_date])
    
    for payment in payments:
        transactions.append({
            'date': payment.sales_payment_date,
            'type': 'Payment',
            'reference': payment.sales_ip_invoice_no.sales_invoice_no,
            'debit': 0,
            'credit': payment.sales_payment_amount
        })
    
    # Sort and calculate balance
    transactions.sort(key=lambda x: x['date'])
    balance = 0
    for trans in transactions:
        balance += trans['debit'] - trans['credit']
        trans['balance'] = balance
    
    total_debit = sum(t['debit'] for t in transactions)
    total_credit = sum(t['credit'] for t in transactions)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer Ledger"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center")
    
    # Title
    ws['A1'] = f"Customer Ledger - {customer.customer_name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    # Customer info
    ws['A2'] = f"Mobile: {customer.customer_mobile}"
    ws.merge_cells('A2:F2')
    
    # Headers
    headers = ['Date', 'Type', 'Invoice No', 'Debit (Sales)', 'Credit (Received)', 'Balance']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Data
    for row, trans in enumerate(transactions, 5):
        ws.cell(row=row, column=1, value=trans['date'].strftime('%d-%m-%Y'))
        ws.cell(row=row, column=2, value=trans['type'])
        ws.cell(row=row, column=3, value=trans['reference'])
        ws.cell(row=row, column=4, value=trans['debit'] if trans['debit'] > 0 else 0)
        ws.cell(row=row, column=5, value=trans['credit'] if trans['credit'] > 0 else 0)
        ws.cell(row=row, column=6, value=trans['balance'])
    
    # Total row
    total_row = len(transactions) + 5
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=total_debit).font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=total_credit).font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=balance).font = Font(bold=True)
    
    # Auto-adjust column widths
    for col_num in range(1, 7):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)
        for row in ws.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                if hasattr(cell, 'value') and cell.value is not None:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="customer_ledger_{customer.customer_name}.xlsx"'
    return response
