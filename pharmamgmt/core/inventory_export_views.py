from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum
from datetime import datetime
import io
import csv

# PDF imports
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Excel imports
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .fast_inventory import FastInventory
from .models import Pharmacy_Details


@login_required
def export_batch_inventory_pdf(request):
    """Export batch-wise inventory report as PDF (Ctrl+Q)"""
    try:
        # Get search query
        search_query = request.GET.get('search', '')
        
        # Get inventory data
        all_inventory_data = FastInventory.get_batch_inventory_data(search_query)
        
        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=landscape(A4), 
            topMargin=0.5*inch, 
            bottomMargin=0.5*inch,
            leftMargin=0.5*inch,
            rightMargin=0.5*inch
        )
        
        # Create styles
        styles = getSampleStyleSheet()
        story = []

        # Get pharmacy details
        try:
            pharmacy = Pharmacy_Details.objects.first()
        except:
            pharmacy = None

        # Title section
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=12,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        )
        
        if pharmacy and pharmacy.pharmaname:
            pharmacy_title = Paragraph(f"{pharmacy.pharmaname}", title_style)
            story.append(pharmacy_title)
        
        report_title = Paragraph("Batch-wise Inventory Report", title_style)
        story.append(report_title)
        
        # Date and search info
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            spaceAfter=6
        )
        
        date_text = Paragraph(f"Generated on: {datetime.now().strftime('%d %B %Y at %H:%M')}", date_style)
        story.append(date_text)
        
        if search_query:
            search_text = Paragraph(f"Search Filter: {search_query}", date_style)
            story.append(search_text)
        
        story.append(Spacer(1, 0.2*inch))

        # Summary Statistics
        total_products = len(all_inventory_data)
        total_value = sum(item['value'] for item in all_inventory_data)
        out_of_stock = sum(1 for item in all_inventory_data if item['stock'] <= 0)
        low_stock = sum(1 for item in all_inventory_data if 0 < item['stock'] <= 10)
        
        summary_data = [
            ['Total Products', 'Total Inventory Value', 'Out of Stock', 'Low Stock (≤10)'],
            [str(total_products), f"₹{total_value:,.2f}", str(out_of_stock), str(low_stock)]
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 2.5*inch, 1.5*inch, 1.5*inch])
        summary_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))

        # Inventory Table
        if all_inventory_data:
            table_data = [['Product Name', 'Company', 'Packing', 'Batch No', 'Expiry', 'Stock Qty', 'MRP', 'Stock Value']]
            
            for item in all_inventory_data:
                # Format expiry date
                expiry_display = item.get('expiry', 'N/A')
                if expiry_display and expiry_display != 'N/A':
                    try:
                        if len(expiry_display) == 7 and '-' in expiry_display:  # MM-YYYY format
                            expiry_display = expiry_display
                        elif len(expiry_display) == 10:  # YYYY-MM-DD format
                            parts = expiry_display.split('-')
                            if len(parts) == 3:
                                expiry_display = f"{parts[1]}-{parts[0]}"
                    except:
                        pass
                
                table_data.append([
                    item['product_name'][:25] + '...' if len(item['product_name']) > 25 else item['product_name'],
                    item['product_company'][:15] + '...' if len(item['product_company']) > 15 else item['product_company'],
                    item['product_packing'][:10] + '...' if len(item['product_packing']) > 10 else item['product_packing'],
                    item['batch_no'][:10] + '...' if len(item['batch_no']) > 10 else item['batch_no'],
                    expiry_display,
                    str(int(item['stock'])),
                    f"₹{item['mrp']:.2f}",
                    f"₹{item['value']:.2f}"
                ])

            # Create table with proper column widths for landscape
            col_widths = [2.2*inch, 1.3*inch, 0.8*inch, 0.8*inch, 0.7*inch, 0.6*inch, 0.7*inch, 0.9*inch]
            inventory_table = Table(table_data, colWidths=col_widths, repeatRows=1)
            
            inventory_table.setStyle(TableStyle([
                # Header style
                ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                
                # Data rows style
                ('FONT', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('ALIGN', (0, 1), (4, -1), 'LEFT'),
                ('ALIGN', (5, 1), (-1, -1), 'RIGHT'),  # Align numeric columns right
                ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
                
                # Grid and alternating colors
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                
                # Highlight zero stock rows
                ('TEXTCOLOR', (5, 1), (5, -1), colors.red),  # Red text for zero stock
            ]))

            story.append(inventory_table)
        else:
            no_data_text = Paragraph("No inventory data found.", styles['Normal'])
            story.append(no_data_text)

        # Build PDF
        doc.build(story)

        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f"batch_inventory_report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        return HttpResponse(f"Error generating PDF: {str(e)}", status=500)


@login_required
def export_batch_inventory_excel(request):
    """Export batch-wise inventory report as Excel (Ctrl+E)"""
    try:
        # Get search query
        search_query = request.GET.get('search', '')
        
        # Get inventory data
        all_inventory_data = FastInventory.get_batch_inventory_data(search_query)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Batch Inventory Report"
        
        # Get pharmacy details
        try:
            pharmacy = Pharmacy_Details.objects.first()
        except:
            pharmacy = None

        # Styles
        title_font = Font(name='Arial', size=14, bold=True, color='1F4E79')
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        data_font = Font(name='Arial', size=10)
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        summary_fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title section
        current_row = 1
        if pharmacy and pharmacy.pharmaname:
            ws.merge_cells(f'A{current_row}:H{current_row}')
            ws[f'A{current_row}'] = pharmacy.pharmaname
            ws[f'A{current_row}'].font = title_font
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
            current_row += 1
        
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = "Batch-wise Inventory Report"
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = f"Generated on: {datetime.now().strftime('%d %B %Y at %H:%M')}"
        ws[f'A{current_row}'].font = Font(name='Arial', size=10)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1
        
        if search_query:
            ws.merge_cells(f'A{current_row}:H{current_row}')
            ws[f'A{current_row}'] = f"Search Filter: {search_query}"
            ws[f'A{current_row}'].font = Font(name='Arial', size=10)
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
            current_row += 1
        
        current_row += 1  # Empty row

        # Summary section
        total_products = len(all_inventory_data)
        total_value = sum(item['value'] for item in all_inventory_data)
        out_of_stock = sum(1 for item in all_inventory_data if item['stock'] <= 0)
        low_stock = sum(1 for item in all_inventory_data if 0 < item['stock'] <= 10)
        
        summary_headers = ['Total Products', 'Total Inventory Value', 'Out of Stock', 'Low Stock (≤10)']
        summary_values = [total_products, f"₹{total_value:,.2f}", out_of_stock, low_stock]
        
        # Summary headers
        for col, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        current_row += 1
        
        # Summary values
        for col, value in enumerate(summary_values, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.font = data_font
            cell.fill = summary_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        current_row += 2  # Empty row
        
        # Data table headers
        headers = ['Product Name', 'Company', 'Packing', 'Batch No', 'Expiry', 'Stock Qty', 'MRP', 'Stock Value']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        current_row += 1
        
        # Data rows
        for item in all_inventory_data:
            # Format expiry date
            expiry_display = item.get('expiry', 'N/A')
            if expiry_display and expiry_display != 'N/A':
                try:
                    if len(expiry_display) == 7 and '-' in expiry_display:  # MM-YYYY format
                        expiry_display = expiry_display
                    elif len(expiry_display) == 10:  # YYYY-MM-DD format
                        parts = expiry_display.split('-')
                        if len(parts) == 3:
                            expiry_display = f"{parts[1]}-{parts[0]}"
                except:
                    pass
            
            row_data = [
                item['product_name'],
                item['product_company'],
                item['product_packing'],
                item['batch_no'],
                expiry_display,
                int(item['stock']),
                f"₹{item['mrp']:.2f}",
                f"₹{item['value']:.2f}"
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.font = data_font
                cell.border = thin_border
                
                # Alignment
                if col in [6, 7, 8]:  # Numeric columns
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')
                
                # Highlight zero stock
                if col == 6 and int(item['stock']) <= 0:
                    cell.font = Font(name='Arial', size=10, color='FF0000', bold=True)
            
            current_row += 1
        
        # Auto-adjust column widths
        column_widths = [30, 20, 15, 15, 12, 12, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"batch_inventory_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return HttpResponse(f"Error generating Excel: {str(e)}", status=500)


@login_required
def export_dateexpiry_inventory_pdf(request):
    """Export date-wise inventory report as PDF (Ctrl+Q)"""
    try:
        # Get search query and filters
        search_query = request.GET.get('search', '')
        expiry_from = request.GET.get('expiry_from', '')
        expiry_to = request.GET.get('expiry_to', '')
        
        # Get inventory data
        expiry_data, total_value = FastInventory.get_dateexpiry_inventory_data(search_query)
        
        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=landscape(A4), 
            topMargin=0.5*inch, 
            bottomMargin=0.5*inch,
            leftMargin=0.5*inch,
            rightMargin=0.5*inch
        )
        
        # Create styles
        styles = getSampleStyleSheet()
        story = []

        # Get pharmacy details
        try:
            pharmacy = Pharmacy_Details.objects.first()
        except:
            pharmacy = None

        # Title section
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=12,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        )
        
        if pharmacy and pharmacy.pharmaname:
            pharmacy_title = Paragraph(f"{pharmacy.pharmaname}", title_style)
            story.append(pharmacy_title)
        
        report_title = Paragraph("Date-wise Inventory Report", title_style)
        story.append(report_title)
        
        # Date and search info
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            spaceAfter=6
        )
        
        date_text = Paragraph(f"Generated on: {datetime.now().strftime('%d %B %Y at %H:%M')}", date_style)
        story.append(date_text)
        
        if search_query:
            search_text = Paragraph(f"Search Filter: {search_query}", date_style)
            story.append(search_text)
        
        story.append(Spacer(1, 0.2*inch))

        # Summary
        total_products = sum(len(group['products']) for group in expiry_data)
        
        summary_data = [
            ['Total Products', 'Total Inventory Value', 'Expiry Groups'],
            [str(total_products), f"₹{total_value:,.2f}", str(len(expiry_data))]
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 2.5*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))

        # Expiry-wise data
        if expiry_data:
            for expiry_group in expiry_data:
                # Group header
                group_style = ParagraphStyle(
                    'GroupHeader',
                    parent=styles['Heading3'],
                    fontSize=12,
                    spaceAfter=6,
                    textColor=colors.darkgreen,
                    backColor=colors.lightgrey
                )
                
                expiry_display = expiry_group['expiry_display']
                group_value = expiry_group['total_value']
                days_info = ""
                
                if expiry_group.get('days_to_expiry') is not None:
                    days = expiry_group['days_to_expiry']
                    if days < 0:
                        days_info = f" (Expired {abs(days)} days ago)"
                    elif days == 0:
                        days_info = " (Expires today)"
                    else:
                        days_info = f" ({days} days remaining)"
                
                group_header = Paragraph(
                    f"{expiry_display}{days_info} - Value: ₹{group_value:,.2f}",
                    group_style
                )
                story.append(group_header)
                
                # Products table for this group
                table_data = [['Product Name', 'Company', 'Packing', 'Batch No', 'Quantity', 'Purchase Rate', 'MRP', 'Stock Value']]
                
                for product in expiry_group['products']:
                    table_data.append([
                        product['product_name'][:25] + '...' if len(product['product_name']) > 25 else product['product_name'],
                        product['product_company'][:15] + '...' if len(product['product_company']) > 15 else product['product_company'],
                        product['product_packing'][:10] + '...' if len(product['product_packing']) > 10 else product['product_packing'],
                        product['batch_no'][:10] + '...' if len(product['batch_no']) > 10 else product['batch_no'],
                        str(int(product['quantity'])),
                        f"₹{product['purchase_rate']:.2f}",
                        f"₹{product['mrp']:.2f}",
                        f"₹{product['value']:.2f}"
                    ])
                
                # Create table
                col_widths = [2.2*inch, 1.3*inch, 0.8*inch, 0.8*inch, 0.6*inch, 0.8*inch, 0.7*inch, 0.9*inch]
                products_table = Table(table_data, colWidths=col_widths, repeatRows=1)
                
                # Determine row color based on expiry status
                row_color = colors.white
                if expiry_group.get('days_to_expiry') is not None:
                    days = expiry_group['days_to_expiry']
                    if days < 0:
                        row_color = colors.mistyrose  # Expired
                    elif days <= 30:
                        row_color = colors.lightyellow  # Expiring soon
                    elif days <= 90:
                        row_color = colors.lightcyan  # Warning
                
                products_table.setStyle(TableStyle([
                    # Header style
                    ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    
                    # Data rows style
                    ('FONT', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('ALIGN', (0, 1), (3, -1), 'LEFT'),
                    ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
                    ('BACKGROUND', (0, 1), (-1, -1), row_color),
                    
                    # Grid
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ]))
                
                story.append(products_table)
                story.append(Spacer(1, 0.2*inch))
        else:
            no_data_text = Paragraph("No inventory data found.", styles['Normal'])
            story.append(no_data_text)

        # Build PDF
        doc.build(story)

        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f"dateexpiry_inventory_report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        return HttpResponse(f"Error generating PDF: {str(e)}", status=500)


@login_required
def export_dateexpiry_inventory_excel(request):
    """Export date-wise inventory report as Excel (Ctrl+E)"""
    try:
        # Get search query and filters
        search_query = request.GET.get('search', '')
        expiry_from = request.GET.get('expiry_from', '')
        expiry_to = request.GET.get('expiry_to', '')
        
        # Get inventory data
        expiry_data, total_value = FastInventory.get_dateexpiry_inventory_data(search_query)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Date-wise Inventory Report"
        
        # Get pharmacy details
        try:
            pharmacy = Pharmacy_Details.objects.first()
        except:
            pharmacy = None

        # Styles
        title_font = Font(name='Arial', size=14, bold=True, color='1F4E79')
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        group_font = Font(name='Arial', size=12, bold=True, color='2F5233')
        data_font = Font(name='Arial', size=10)
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        summary_fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
        group_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        expired_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
        expiring_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title section
        current_row = 1
        if pharmacy and pharmacy.pharmaname:
            ws.merge_cells(f'A{current_row}:H{current_row}')
            ws[f'A{current_row}'] = pharmacy.pharmaname
            ws[f'A{current_row}'].font = title_font
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
            current_row += 1
        
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = "Date-wise Inventory Report"
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = f"Generated on: {datetime.now().strftime('%d %B %Y at %H:%M')}"
        ws[f'A{current_row}'].font = Font(name='Arial', size=10)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1
        
        if search_query:
            ws.merge_cells(f'A{current_row}:H{current_row}')
            ws[f'A{current_row}'] = f"Search Filter: {search_query}"
            ws[f'A{current_row}'].font = Font(name='Arial', size=10)
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
            current_row += 1
        
        current_row += 1  # Empty row

        # Summary section
        total_products = sum(len(group['products']) for group in expiry_data)
        
        summary_headers = ['Total Products', 'Total Inventory Value', 'Expiry Groups']
        summary_values = [total_products, f"₹{total_value:,.2f}", len(expiry_data)]
        
        # Summary headers
        for col, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        current_row += 1
        
        # Summary values
        for col, value in enumerate(summary_values, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.font = data_font
            cell.fill = summary_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        current_row += 2  # Empty row
        
        # Data headers
        headers = ['Expiry Date', 'Product Name', 'Company', 'Packing', 'Batch No', 'Quantity', 'Purchase Rate', 'MRP', 'Stock Value']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        current_row += 1
        
        # Data rows grouped by expiry
        for expiry_group in expiry_data:
            # Group header row
            expiry_display = expiry_group['expiry_display']
            group_value = expiry_group['total_value']
            days_info = ""
            
            if expiry_group.get('days_to_expiry') is not None:
                days = expiry_group['days_to_expiry']
                if days < 0:
                    days_info = f" (Expired {abs(days)} days ago)"
                elif days == 0:
                    days_info = " (Expires today)"
                else:
                    days_info = f" ({days} days remaining)"
            
            # Merge cells for group header
            ws.merge_cells(f'A{current_row}:H{current_row}')
            group_cell = ws[f'A{current_row}']
            group_cell.value = f"{expiry_display}{days_info} - Value: ₹{group_value:,.2f}"
            group_cell.font = group_font
            group_cell.fill = group_fill
            group_cell.alignment = Alignment(horizontal='left')
            group_cell.border = thin_border
            
            # Value cell
            value_cell = ws.cell(row=current_row, column=9, value=f"₹{group_value:,.2f}")
            value_cell.font = group_font
            value_cell.fill = group_fill
            value_cell.alignment = Alignment(horizontal='right')
            value_cell.border = thin_border
            
            current_row += 1
            
            # Determine fill color based on expiry status
            row_fill = None
            if expiry_group.get('days_to_expiry') is not None:
                days = expiry_group['days_to_expiry']
                if days < 0:
                    row_fill = expired_fill  # Expired
                elif days <= 30:
                    row_fill = expiring_fill  # Expiring soon
            
            # Products in this group
            for product in expiry_group['products']:
                row_data = [
                    '',  # Empty for expiry date (already shown in group header)
                    product['product_name'],
                    product['product_company'],
                    product['product_packing'],
                    product['batch_no'],
                    int(product['quantity']),
                    f"₹{product['purchase_rate']:.2f}",
                    f"₹{product['mrp']:.2f}",
                    f"₹{product['value']:.2f}"
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col, value=value)
                    cell.font = data_font
                    cell.border = thin_border
                    
                    if row_fill:
                        cell.fill = row_fill
                    
                    # Alignment
                    if col in [6, 7, 8, 9]:  # Numeric columns
                        cell.alignment = Alignment(horizontal='right')
                    else:
                        cell.alignment = Alignment(horizontal='left')
                
                current_row += 1
            
            current_row += 1  # Empty row between groups
        
        # Auto-adjust column widths
        column_widths = [15, 30, 20, 15, 15, 12, 15, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"dateexpiry_inventory_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return HttpResponse(f"Error generating Excel: {str(e)}", status=500)