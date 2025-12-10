from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Q
from django.core.paginator import Paginator
from django.template.loader import render_to_string
from .models import SalesMaster, CustomerMaster, SalesInvoiceMaster, CustomerChallanMaster
from datetime import datetime, date
import io
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

@login_required
def customer_wise_sales_report(request):
    """Customer-wise sales report with date filters"""
    
    # Get filter parameters
    customer_id = request.GET.get('customer_id')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    # Initialize context
    context = {
        'title': 'Customer Wise Sales Report',
        'customers': CustomerMaster.objects.all().order_by('customer_name'),
        'selected_customer_id': customer_id,
        'from_date': from_date,
        'to_date': to_date,
        'sales_data': [],
        'summary': {}
    }
    
    # Handle export requests
    export_type = request.GET.get('export')
    if export_type and customer_id and from_date and to_date:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or export_type == 'print':
            try:
                customer = CustomerMaster.objects.get(customerid=customer_id)
                from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
                to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()
                
                # Get sales data for export
                sales_invoices = SalesInvoiceMaster.objects.filter(
                    customerid=customer,
                    sales_invoice_date__range=[from_date_obj, to_date_obj]
                ).order_by('-sales_invoice_date')
                
                sales_data = []
                for invoice in sales_invoices:
                    invoice_items = SalesMaster.objects.filter(sales_invoice_no=invoice).select_related('productid')
                    invoice_total = sum(item.sale_total_amount for item in invoice_items)
                    sales_data.append({
                        'invoice': invoice,
                        'items': invoice_items,
                        'invoice_total': invoice_total
                    })
                
                # Get customer challans
                customer_challans = list(CustomerChallanMaster.objects.filter(
                    customer_name=customer,
                    sales_entry_date__date__range=[from_date_obj, to_date_obj]
                ).select_related('product_id'))
                
                if export_type == 'pdf':
                    return export_customer_sales_pdf(customer, sales_data, customer_challans, from_date_obj, to_date_obj)
                elif export_type == 'print':
                    # Render print view HTML
                    from .models import Pharmacy_Details
                    pharmacy = Pharmacy_Details.objects.first()
                    context = {
                        'pharmacy': pharmacy,
                        'customer': customer,
                        'sales_data': sales_data,
                        'customer_challans': customer_challans,
                        'from_date': from_date_obj,
                        'to_date': to_date_obj,
                        'summary': {
                            'total_amount': sum(s['invoice_total'] for s in sales_data) + sum(c.sale_total_amount for c in customer_challans),
                            'total_qty': sum(sum(i.sale_quantity for i in s['items']) for s in sales_data) + sum(c.sale_quantity for c in customer_challans)
                        }
                    }
                    return render(request, 'reports/customer_sales_print.html', context)
                elif export_type == 'excel':
                    return export_customer_sales_excel(customer, sales_data, customer_challans, from_date_obj, to_date_obj)
            except Exception as e:
                import traceback
                print(f"Export error: {str(e)}")
                print(traceback.format_exc())
                return JsonResponse({'success': False, 'error': f'Export failed: {str(e)}'})
    
    # If filters applied, get sales data
    if customer_id and from_date and to_date:
        try:
            customer = CustomerMaster.objects.get(customerid=customer_id)
            
            # Parse dates
            from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
            to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()
            
            # Get sales invoices
            sales_invoices = SalesInvoiceMaster.objects.filter(
                customerid=customer,
                sales_invoice_date__range=[from_date_obj, to_date_obj]
            ).order_by('-sales_invoice_date')
            
            # Get sales data
            sales_data = []
            total_amount = 0
            total_quantity = 0
            
            for invoice in sales_invoices:
                # Get invoice items
                invoice_items = SalesMaster.objects.filter(
                    sales_invoice_no=invoice
                ).select_related('productid')
                
                invoice_total = sum(item.sale_total_amount for item in invoice_items)
                invoice_qty = sum(item.sale_quantity for item in invoice_items)
                
                sales_data.append({
                    'invoice': invoice,
                    'items': invoice_items,
                    'invoice_total': invoice_total,
                    'invoice_qty': invoice_qty,
                    'balance_due': invoice.balance_due
                })
                
                total_amount += invoice_total
                total_quantity += invoice_qty
            
            # Get customer challan sales
            customer_challans = CustomerChallanMaster.objects.filter(
                customer_name=customer,
                sales_entry_date__date__range=[from_date_obj, to_date_obj]
            ).select_related('product_id')
            
            challan_total = sum(item.sale_total_amount for item in customer_challans)
            challan_qty = sum(item.sale_quantity for item in customer_challans)
            
            context.update({
                'selected_customer': customer,
                'sales_data': sales_data,
                'customer_challans': customer_challans,
                'summary': {
                    'total_invoices': len(sales_data),
                    'total_amount': total_amount,
                    'total_quantity': total_quantity,
                    'challan_amount': challan_total,
                    'challan_quantity': challan_qty,
                    'grand_total': total_amount + challan_total,
                    'grand_quantity': total_quantity + challan_qty
                }
            })
            
        except CustomerMaster.DoesNotExist:
            context['error'] = 'Customer not found'
        except ValueError:
            context['error'] = 'Invalid date format'
    
    return render(request, 'reports/customer_wise_sales.html', context)

@login_required
def quick_customer_search(request):
    """Quick customer search API for Ctrl+F2 functionality"""
    
    search_term = request.GET.get('q', '').strip()
    
    if len(search_term) < 2:
        return JsonResponse({
            'success': True,
            'customers': [],
            'message': 'Type at least 2 characters'
        })
    
    # Search customers by name, mobile, or GST
    customers = CustomerMaster.objects.filter(
        Q(customer_name__icontains=search_term) |
        Q(customer_mobile__icontains=search_term) |
        Q(customer_gstno__icontains=search_term)
    ).order_by('customer_name')[:20]  # Limit to 20 results
    
    customer_data = []
    for customer in customers:
        customer_data.append({
            'customerid': customer.customerid,
            'customer_name': customer.customer_name,
            'customer_mobile': customer.customer_mobile,
            'customer_gstno': customer.customer_gstno,
            'customer_type': customer.customer_type
        })
    
    return JsonResponse({
        'success': True,
        'customers': customer_data
    })

@login_required
def customer_sales_summary(request, customer_id):
    """Get customer sales summary for quick view"""
    
    try:
        customer = CustomerMaster.objects.get(customerid=customer_id)
        
        # Get last 30 days sales
        from datetime import timedelta
        end_date = date.today()
        start_date = end_date - timedelta(days=30)
        
        # Invoice sales
        recent_sales = SalesMaster.objects.filter(
            customerid=customer,
            sale_entry_date__date__range=[start_date, end_date]
        ).aggregate(
            total_amount=Sum('sale_total_amount'),
            total_quantity=Sum('sale_quantity')
        )
        
        # Challan sales
        recent_challans = CustomerChallanMaster.objects.filter(
            customer_name=customer,
            sales_entry_date__date__range=[start_date, end_date]
        ).aggregate(
            total_amount=Sum('sale_total_amount'),
            total_quantity=Sum('sale_quantity')
        )
        
        # Outstanding balance
        outstanding_invoices = SalesInvoiceMaster.objects.filter(
            customerid=customer
        )
        
        total_outstanding = sum(invoice.balance_due for invoice in outstanding_invoices)
        
        return JsonResponse({
            'success': True,
            'customer_name': customer.customer_name,
            'recent_sales': {
                'amount': (recent_sales['total_amount'] or 0) + (recent_challans['total_amount'] or 0),
                'quantity': (recent_sales['total_quantity'] or 0) + (recent_challans['total_quantity'] or 0)
            },
            'outstanding_balance': total_outstanding,
            'customer_type': customer.customer_type
        })
        
    except CustomerMaster.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Customer not found'
        })

def export_customer_sales_pdf(customer, sales_data, customer_challans, from_date, to_date):
    """Export customer sales to professional PDF document"""
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.units import mm
    from datetime import datetime
    
    response = HttpResponse(content_type='application/pdf')
    filename = f'customer_sales_{customer.customer_name}_{from_date.strftime("%Y%m%d")}_{to_date.strftime("%Y%m%d")}.pdf'
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    
    # A4 Portrait with professional margins
    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=20*mm, leftMargin=20*mm, topMargin=25*mm, bottomMargin=20*mm)
    elements = []
    
    styles = getSampleStyleSheet()
    
    # Professional Header Style
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.darkblue,
        alignment=TA_CENTER,
        spaceAfter=8*mm,
        fontName='Helvetica-Bold'
    )
    
    # Company/Report Title
    try:
        from .models import Pharmacy_Details
        pharmacy = Pharmacy_Details.objects.first()
        if pharmacy and pharmacy.pharmaname:
            company_title = Paragraph(pharmacy.pharmaname.upper(), header_style)
            elements.append(company_title)
    except:
        pass
    
    # Report Title
    report_title = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.black,
        alignment=TA_CENTER,
        spaceAfter=6*mm,
        fontName='Helvetica-Bold'
    )
    elements.append(Paragraph("CUSTOMER SALES REPORT", report_title))
    
    # Customer Info Section
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.black,
        alignment=TA_LEFT,
        spaceAfter=3*mm,
        fontName='Helvetica'
    )
    
    # Customer Details Box
    customer_info = [
        ['Customer Name:', customer.customer_name],
        ['Customer Type:', customer.customer_type or 'N/A'],
        ['Mobile:', customer.customer_mobile or 'N/A'],
        ['GST No:', customer.customer_gstno or 'N/A'],
        ['Report Period:', f"{from_date.strftime('%d/%m/%Y')} to {to_date.strftime('%d/%m/%Y')}"],
        ['Generated On:', datetime.now().strftime('%d/%m/%Y %H:%M')]
    ]
    
    info_table = Table(customer_info, colWidths=[40*mm, 80*mm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4)
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 8*mm))
    
    # Sales Data Table
    if sales_data or customer_challans:
        # Table Headers
        data = [['S.No.', 'Type', 'Invoice/Challan', 'Date', 'Product', 'Batch', 'Qty', 'Rate', 'Amount']]
        
        total_amount = 0
        total_qty = 0
        sr_no = 1
        
        # Add invoice sales
        for sale_info in sales_data:
            invoice = sale_info['invoice']
            for item in sale_info['items']:
                data.append([
                    str(sr_no),
                    'Invoice',
                    str(invoice.sales_invoice_no),
                    invoice.sales_invoice_date.strftime('%d/%m/%Y'),
                    item.product_name[:20] + '...' if len(item.product_name) > 20 else item.product_name,
                    item.product_batch_no[:8] if item.product_batch_no else 'N/A',
                    f"{item.sale_quantity:.0f}",
                    f"{item.sale_rate:.2f}",
                    f"{item.sale_total_amount:.2f}"
                ])
                total_amount += item.sale_total_amount
                total_qty += item.sale_quantity
                sr_no += 1
        
        # Add challan sales
        for challan in customer_challans:
            data.append([
                str(sr_no),
                'Challan',
                str(challan.customer_challan_no),
                challan.sales_entry_date.strftime('%d/%m/%Y'),
                challan.product_name[:20] + '...' if len(challan.product_name) > 20 else challan.product_name,
                challan.product_batch_no[:8] if challan.product_batch_no else 'N/A',
                f"{challan.sale_quantity:.0f}",
                f"{challan.sale_rate:.2f}",
                f"{challan.sale_total_amount:.2f}"
            ])
            total_amount += challan.sale_total_amount
            total_qty += challan.sale_quantity
            sr_no += 1
        
        # Summary Row
        data.append(['', '', '', '', 'TOTAL', '', f"{total_qty:.0f}", '', f"{total_amount:.2f}"])
        
        # Create table with proper column widths
        col_widths = [10*mm, 15*mm, 22*mm, 18*mm, 42*mm, 16*mm, 12*mm, 18*mm, 22*mm]
        sales_table = Table(data, colWidths=col_widths, repeatRows=1)
        
        sales_table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Data rows styling
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 8),
            ('ALIGN', (0, 1), (0, -2), 'CENTER'),  # S.No.
            ('ALIGN', (1, 1), (1, -2), 'CENTER'),  # Type
            ('ALIGN', (2, 1), (2, -2), 'LEFT'),    # Invoice
            ('ALIGN', (3, 1), (3, -2), 'CENTER'),  # Date
            ('ALIGN', (4, 1), (4, -2), 'LEFT'),    # Product
            ('ALIGN', (5, 1), (5, -2), 'CENTER'),  # Batch
            ('ALIGN', (6, 1), (-1, -2), 'RIGHT'),  # Qty, Rate, Amount
            
            # Total row styling
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 9),
            ('ALIGN', (0, -1), (5, -1), 'CENTER'),
            ('ALIGN', (6, -1), (-1, -1), 'RIGHT'),
            
            # Grid and borders
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.lightgrey]),
            
            # Padding
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4)
        ]))
        
        elements.append(sales_table)
        
        # Summary Section
        elements.append(Spacer(1, 8*mm))
        
        summary_data = [
            ['Total Invoices:', str(len(sales_data))],
            ['Total Items:', str(sr_no - 1)],
            ['Total Quantity:', f"{total_qty:.0f}"],
            ['Total Amount:', f"₹ {total_amount:,.2f}"]
        ]
        
        summary_table = Table(summary_data, colWidths=[40*mm, 40*mm])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightblue),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4)
        ]))
        
        elements.append(summary_table)
    else:
        no_data = Paragraph("No sales data found for the selected period.", info_style)
        elements.append(no_data)
    
    # Footer
    elements.append(Spacer(1, 10*mm))
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    footer_text = f"Report generated on {datetime.now().strftime('%d/%m/%Y at %H:%M')} | Page 1"
    elements.append(Paragraph(footer_text, footer_style))
    
    doc.build(elements)
    return response

def export_customer_sales_excel(customer, sales_data, customer_challans, from_date, to_date):
    """Export customer sales to Excel"""
    from io import BytesIO
    
    # Create workbook in memory
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Customer Sales'
    
    # Fetch pharmacy details from database
    row = 1
    try:
        from .models import Pharmacy_Details
        pharmacy = Pharmacy_Details.objects.first()
        if pharmacy:
            # Pharmacy Name
            if pharmacy.pharmaname:
                cell = worksheet.cell(row=row, column=1, value=pharmacy.pharmaname)
                cell.font = Font(bold=True, size=14)
                cell.alignment = Alignment(horizontal='center')
                worksheet.merge_cells(f'A{row}:H{row}')
                row += 1
            
            # Proprietor Name
            if pharmacy.proprietorname:
                cell = worksheet.cell(row=row, column=1, value=f"Proprietor: {pharmacy.proprietorname}")
                cell.font = Font(size=11)
                cell.alignment = Alignment(horizontal='center')
                worksheet.merge_cells(f'A{row}:H{row}')
                row += 1
            
            # Contact Details
            if pharmacy.proprietorcontact:
                cell = worksheet.cell(row=row, column=1, value=f"Contact: {pharmacy.proprietorcontact}")
                cell.font = Font(size=10)
                cell.alignment = Alignment(horizontal='center')
                worksheet.merge_cells(f'A{row}:H{row}')
                row += 1
            
            # Email
            if pharmacy.proprietoremail:
                cell = worksheet.cell(row=row, column=1, value=f"Email: {pharmacy.proprietoremail}")
                cell.font = Font(size=10)
                cell.alignment = Alignment(horizontal='center')
                worksheet.merge_cells(f'A{row}:H{row}')
                row += 1
            
            # Website URL
            if pharmacy.pharmaweburl:
                cell = worksheet.cell(row=row, column=1, value=f"Website: {pharmacy.pharmaweburl}")
                cell.font = Font(size=10)
                cell.alignment = Alignment(horizontal='center')
                worksheet.merge_cells(f'A{row}:H{row}')
                row += 1
            
            # Empty row for spacing
            row += 1
    except Exception as e:
        print(f"Error fetching pharmacy details: {e}")
    
    # Report Title
    cell = worksheet.cell(row=row, column=1, value='CUSTOMER WISE SALES REPORT')
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')
    worksheet.merge_cells(f'A{row}:H{row}')
    row += 1
    
    # Customer and Date Info
    cell = worksheet.cell(row=row, column=1, value=f"Customer: {customer.customer_name}")
    cell.font = Font(bold=True, size=11)
    worksheet.merge_cells(f'A{row}:D{row}')
    cell = worksheet.cell(row=row, column=5, value=f"Period: {from_date.strftime('%d/%m/%Y')} to {to_date.strftime('%d/%m/%Y')}")
    cell.font = Font(bold=True, size=11)
    worksheet.merge_cells(f'E{row}:H{row}')
    row += 1
    
    # Empty row
    row += 1
    
    # Column Headers
    headers = ['Type', 'Invoice/Challan', 'Date', 'Product Name', 'Batch', 'Quantity', 'Rate', 'Amount']
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    # Data
    total_amount = 0
    
    # Invoice sales
    for sale_info in sales_data:
        invoice = sale_info['invoice']
        for item in sale_info['items']:
            worksheet.cell(row=row, column=1, value='Invoice')
            worksheet.cell(row=row, column=2, value=str(invoice.sales_invoice_no))
            worksheet.cell(row=row, column=3, value=invoice.sales_invoice_date.strftime('%d/%m/%Y'))
            worksheet.cell(row=row, column=4, value=item.product_name)
            worksheet.cell(row=row, column=5, value=item.product_batch_no)
            worksheet.cell(row=row, column=6, value=item.sale_quantity)
            worksheet.cell(row=row, column=7, value=item.sale_rate)
            worksheet.cell(row=row, column=8, value=item.sale_total_amount)
            total_amount += item.sale_total_amount
            row += 1
    
    # Challan sales
    for challan in customer_challans:
        worksheet.cell(row=row, column=1, value='Challan')
        worksheet.cell(row=row, column=2, value=str(challan.customer_challan_no))
        worksheet.cell(row=row, column=3, value=challan.sales_entry_date.strftime('%d/%m/%Y'))
        worksheet.cell(row=row, column=4, value=challan.product_name)
        worksheet.cell(row=row, column=5, value=challan.product_batch_no)
        worksheet.cell(row=row, column=6, value=challan.sale_quantity)
        worksheet.cell(row=row, column=7, value=challan.sale_rate)
        worksheet.cell(row=row, column=8, value=challan.sale_total_amount)
        total_amount += challan.sale_total_amount
        row += 1
    
    # Total row
    worksheet.cell(row=row, column=7, value='Total:').font = Font(bold=True)
    worksheet.cell(row=row, column=8, value=total_amount).font = Font(bold=True)
    
    # Auto-adjust columns
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, 9):  # Columns A to H (1 to 8)
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row_cells in worksheet.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row_cells:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO buffer
    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create response with proper headers
    response = HttpResponse(
        excel_buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'customer_sales_{customer.customer_name}_{from_date.strftime("%Y%m%d")}_{to_date.strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['Content-Length'] = excel_buffer.tell()
    
    return response

