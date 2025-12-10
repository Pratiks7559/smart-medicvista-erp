from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import HttpResponse
from django.template.loader import get_template
from django.core.paginator import Paginator
from datetime import datetime
import pandas as pd
import io
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from .models import InvoiceMaster, PurchaseMaster, Pharmacy_Details

@login_required
def purchase2_report(request):
    """Simple Marg-style purchase report with from-to date filter and pagination"""
    
    # Get date range from request
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    
    # Parse dates
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else None
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else None
    except ValueError:
        start_date = None
        end_date = None
    
    # Get purchase invoices - always show all, filter only if dates provided
    invoices = InvoiceMaster.objects.all().order_by('-invoice_date')
    
    # Apply date filter only if both dates provided
    if start_date and end_date:
        invoices = invoices.filter(invoice_date__range=[start_date, end_date])
    
    # Calculate totals for each invoice
    invoice_data = []
    total_purchases = 0
    total_paid = 0
    total_pending = 0
    
    for invoice in invoices:
        # Calculate invoice total from purchase items
        invoice_total = PurchaseMaster.objects.filter(
            product_invoiceid=invoice.invoiceid
        ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        
        balance = invoice_total - invoice.invoice_paid
        
        invoice_data.append({
            'invoice_no': invoice.invoice_no,
            'date': invoice.invoice_date,
            'supplier_name': invoice.supplierid.supplier_name,
            'total': invoice_total,
            'paid': invoice.invoice_paid,
            'balance': balance
        })
        
        total_purchases += invoice_total
        total_paid += invoice.invoice_paid
        total_pending += balance
    
    # Pagination - 15 invoices per page
    paginator = Paginator(invoice_data, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get pharmacy details
    pharmacy = Pharmacy_Details.objects.first()
    
    context = {
        'title': 'Purchase Report',
        'start_date': start_date,
        'end_date': end_date,
        'page_obj': page_obj,
        'total_purchases': total_purchases,
        'total_paid': total_paid,
        'total_pending': total_pending,
        'pharmacy': pharmacy,
        'has_filter': bool(start_date and end_date),
    }
    
    return render(request, 'reports/purchase2_report.html', context)

@login_required
def purchase2_report_pdf(request):
    """Export purchase report as PDF"""
    # Get same data as main report
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else None
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else None
    except ValueError:
        start_date = None
        end_date = None
    
    # Get data - all invoices or filtered
    invoices = InvoiceMaster.objects.all().order_by('-invoice_date')
    if start_date and end_date:
        invoices = invoices.filter(invoice_date__range=[start_date, end_date])
    
    invoice_data = []
    total_purchases = total_paid = total_pending = 0
    
    for invoice in invoices:
        invoice_total = PurchaseMaster.objects.filter(
            product_invoiceid=invoice.invoiceid
        ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        
        balance = invoice_total - invoice.invoice_paid
        invoice_data.append({
            'invoice_no': invoice.invoice_no,
            'date': invoice.invoice_date.strftime('%d-%m-%Y'),
            'supplier_name': invoice.supplierid.supplier_name,
            'total': invoice_total,
            'paid': invoice.invoice_paid,
            'balance': balance
        })
        
        total_purchases += invoice_total
        total_paid += invoice.invoice_paid
        total_pending += balance
    
    # Create PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="purchase_report_{start_date}_{end_date}.pdf"'
    response['X-Frame-Options'] = 'SAMEORIGIN'
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
    
    # Get pharmacy details
    pharmacy = Pharmacy_Details.objects.first()
    
    # Build PDF content
    story = []
    styles = getSampleStyleSheet()
    
    # Pharmacy Header
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], alignment=1, fontSize=16)
    info_style = ParagraphStyle('InfoStyle', parent=styles['Normal'], alignment=1, fontSize=9)
    
    if pharmacy:
        if pharmacy.pharmaname:
            story.append(Paragraph(pharmacy.pharmaname.upper(), title_style))
        pharmacy_info = []
        if pharmacy.proprietorname:
            pharmacy_info.append(f"Proprietor: {pharmacy.proprietorname}")
        if pharmacy.proprietorcontact:
            pharmacy_info.append(f"Contact: {pharmacy.proprietorcontact}")
        if pharmacy.proprietoremail:
            pharmacy_info.append(f"Email: {pharmacy.proprietoremail}")
        if pharmacy.pharmaweburl:
            pharmacy_info.append(f"Website: {pharmacy.pharmaweburl}")
        if pharmacy_info:
            story.append(Paragraph(" | ".join(pharmacy_info), info_style))
    story.append(Spacer(1, 12))
    
    # Report title
    story.append(Paragraph('PURCHASE REPORT', title_style))
    story.append(Paragraph(f'Period: {start_date.strftime("%d-%m-%Y")} to {end_date.strftime("%d-%m-%Y")}', styles['Normal']))
    story.append(Spacer(1, 12))
    
    # Table data
    data = [['Invoice No', 'Date', 'Supplier', 'Total (₹)', 'Paid (₹)', 'Balance (₹)']]
    for item in invoice_data:
        data.append([
            item['invoice_no'], item['date'], item['supplier_name'][:20],
            f"{item['total']:.2f}", f"{item['paid']:.2f}", f"{item['balance']:.2f}"
        ])
    
    # Add totals
    data.append(['', '', 'TOTAL', f'{total_purchases:.2f}', f'{total_paid:.2f}', f'{total_pending:.2f}'])
    
    # Create table
    table = Table(data, colWidths=[1.2*inch, 1*inch, 2*inch, 1*inch, 1*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    doc.build(story)
    
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response

@login_required
def purchase2_report_excel(request):
    """Export purchase report as Excel"""
    # Get same data as main report
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else None
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else None
    except ValueError:
        start_date = None
        end_date = None
    
    # Get data - all invoices or filtered
    invoices = InvoiceMaster.objects.all().order_by('-invoice_date')
    if start_date and end_date:
        invoices = invoices.filter(invoice_date__range=[start_date, end_date])
    
    data = []
    total_purchases = total_paid = total_pending = 0
    
    for invoice in invoices:
        invoice_total = PurchaseMaster.objects.filter(
            product_invoiceid=invoice.invoiceid
        ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        
        balance = invoice_total - invoice.invoice_paid
        data.append({
            'Invoice No': invoice.invoice_no,
            'Date': invoice.invoice_date.strftime('%d-%m-%Y'),
            'Supplier': invoice.supplierid.supplier_name,
            'Total': invoice_total,
            'Paid': invoice.invoice_paid,
            'Balance': balance
        })
        
        total_purchases += invoice_total
        total_paid += invoice.invoice_paid
        total_pending += balance
    
    # Add totals row
    data.append({
        'Invoice No': '',
        'Date': '',
        'Supplier': 'TOTAL',
        'Total': total_purchases,
        'Paid': total_paid,
        'Balance': total_pending
    })
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Create Excel response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="purchase_report_{start_date}_{end_date}.xlsx"'
    
    # Write to Excel with enhanced formatting
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Purchase Report', index=False, startrow=4)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Purchase Report']
        
        # Import styling modules
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        
        # Get pharmacy details for header
        pharmacy = Pharmacy_Details.objects.first()
        
        # Add pharmacy header
        row = 1
        if pharmacy:
            if pharmacy.pharmaname:
                worksheet[f'A{row}'] = pharmacy.pharmaname.upper()
                worksheet[f'A{row}'].font = Font(size=16, bold=True)
                worksheet[f'A{row}'].alignment = Alignment(horizontal='center')
                worksheet.merge_cells(f'A{row}:F{row}')
                row += 1
            
            pharmacy_info = []
            if pharmacy.proprietorname:
                pharmacy_info.append(f"Proprietor: {pharmacy.proprietorname}")
            if pharmacy.proprietorcontact:
                pharmacy_info.append(f"Contact: {pharmacy.proprietorcontact}")
            if pharmacy.proprietoremail:
                pharmacy_info.append(f"Email: {pharmacy.proprietoremail}")
            if pharmacy.pharmaweburl:
                pharmacy_info.append(f"Website: {pharmacy.pharmaweburl}")
            
            if pharmacy_info:
                worksheet[f'A{row}'] = " | ".join(pharmacy_info)
                worksheet[f'A{row}'].font = Font(size=10)
                worksheet[f'A{row}'].alignment = Alignment(horizontal='center')
                worksheet.merge_cells(f'A{row}:F{row}')
                row += 1
        
        worksheet['A3'] = f'PURCHASE REPORT - Period: {start_date.strftime("%d-%m-%Y")} to {end_date.strftime("%d-%m-%Y")}'
        worksheet['A3'].font = Font(size=12, bold=True)
        worksheet['A3'].alignment = Alignment(horizontal='center')
        worksheet.merge_cells('A3:F3')
        
        # Style header row
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header styling
        for col in range(1, 7):  # A to F columns
            cell = worksheet.cell(row=5, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Style data rows
        data_fill_1 = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        data_fill_2 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        for row in range(6, len(data) + 5):  # Data rows
            fill = data_fill_1 if row % 2 == 0 else data_fill_2
            for col in range(1, 7):
                cell = worksheet.cell(row=row, column=col)
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center' if col <= 3 else 'right', vertical='center')
                
                # Format currency columns
                if col >= 4:  # Total, Paid, Balance columns
                    cell.number_format = '₹#,##0.00'
        
        # Style totals row
        total_row = len(data) + 4
        total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        total_font = Font(bold=True)
        
        for col in range(1, 7):
            cell = worksheet.cell(row=total_row, column=col)
            cell.fill = total_fill
            cell.font = total_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col <= 3 else 'right', vertical='center')
            if col >= 4:
                cell.number_format = '₹#,##0.00'
        
        # Auto-adjust column widths
        column_widths = [15, 12, 25, 15, 15, 15]  # Custom widths
        for i, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width
        
        # Add borders to header section
        for row in range(1, 4):
            for col in range(1, 7):
                worksheet.cell(row=row, column=col).border = border
    
    return response