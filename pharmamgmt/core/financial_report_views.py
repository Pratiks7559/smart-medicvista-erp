from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count, Q, F, Case, When, DecimalField, FloatField
from django.db.models.functions import TruncMonth, TruncDate, Coalesce, Cast
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import json
import csv
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference

from .models import (
    SalesInvoiceMaster, SalesMaster, SalesInvoicePaid,
    InvoiceMaster, PurchaseMaster, InvoicePaid,
    ReturnSalesInvoiceMaster, ReturnSalesMaster, ReturnSalesInvoicePaid,
    ReturnInvoiceMaster, ReturnPurchaseMaster, PurchaseReturnInvoicePaid,
    CustomerMaster, SupplierMaster, ProductMaster
)

@login_required
def financial_report(request):
    """Comprehensive financial report with all key metrics"""
    
    # Get date range from request
    end_date = request.GET.get('end_date')
    start_date = request.GET.get('start_date')
    
    if not end_date:
        end_date = timezone.now().date()
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    if not start_date:
        start_date = end_date - timedelta(days=30)
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    
    # Sales Data
    sales_data = get_sales_data(start_date, end_date)
    
    # Purchase Data
    purchase_data = get_purchase_data(start_date, end_date)
    
    # Returns Data
    returns_data = get_returns_data(start_date, end_date)
    
    # Outstanding Amounts
    outstanding_data = get_outstanding_amounts()
    
    # Monthly Trends
    monthly_trends = get_monthly_trends()
    
    # Inventory Metrics
    inventory_metrics = get_inventory_metrics()
    
    # Profit Analysis
    profit_analysis = get_profit_analysis(start_date, end_date)
    
    # Cash Flow Data
    cash_flow_data = get_cash_flow_data(start_date, end_date)
    
    # Top Performers
    top_performers = get_top_performers(start_date, end_date)
    
    # Calculate key metrics
    gross_sales = sales_data['total_sales']
    gross_purchases = purchase_data['total_purchases']
    sales_returns = returns_data['sales_returns']
    purchase_returns = returns_data['purchase_returns']
    
    net_sales = gross_sales - sales_returns
    net_purchases = gross_purchases - purchase_returns
    gross_profit = net_sales - net_purchases
    
    context = {
        'title': 'Financial Report',
        'start_date': start_date,
        'end_date': end_date,
        
        # Basic Metrics
        'sales': gross_sales,
        'purchases': gross_purchases,
        'sales_returns': sales_returns,
        'purchase_returns': purchase_returns,
        'net_sales': net_sales,
        'net_purchases': net_purchases,
        'gross_profit': gross_profit,
        
        # Detailed Data
        'sales_data': sales_data,
        'purchase_data': purchase_data,
        'returns_data': returns_data,
        'outstanding_receivables': outstanding_data['receivables'],
        'outstanding_payables': outstanding_data['payables'],
        'total_receivables': outstanding_data['total_receivables'],
        'total_payables': outstanding_data['total_payables'],
        
        # Trends and Analytics
        'monthly_sales': monthly_trends['sales'],
        'monthly_purchases': monthly_trends['purchases'],
        'inventory_metrics': inventory_metrics,
        'profit_analysis': profit_analysis,
        'cash_flow_data': cash_flow_data,
        'top_performers': top_performers,
        
        # Additional Metrics
        'profit_margin': (gross_profit / net_sales * 100) if net_sales > 0 else 0,
        'inventory_turnover': inventory_metrics.get('turnover_ratio', 0),
        'days_sales_outstanding': outstanding_data.get('dso', 0),
        'days_payable_outstanding': outstanding_data.get('dpo', 0),
    }
    
    return render(request, 'reports/financial_report.html', context)

def get_sales_data(start_date, end_date):
    """Get comprehensive sales data"""
    
    try:
        # Regular Sales Invoices
        regular_sales_total = SalesMaster.objects.filter(
            sale_entry_date__date__range=[start_date, end_date]
        ).aggregate(total=Sum('sale_total_amount'))['total'] or 0
        
        regular_sales_count = SalesMaster.objects.filter(
            sale_entry_date__date__range=[start_date, end_date]
        ).count()
        
        # Payments Received
        payments_received = SalesInvoicePaid.objects.filter(
            sales_payment_date__range=[start_date, end_date]
        ).aggregate(total=Sum('sales_payment_amount'))['total'] or 0
        
        total_sales = float(regular_sales_total)
        total_payments = float(payments_received)
        
        return {
            'total_sales': total_sales,
            'regular_sales': total_sales,
            'challan_sales': 0,
            'invoice_count': regular_sales_count,
            'avg_invoice_value': total_sales / max(1, regular_sales_count),
            'payments_received': total_payments,
            'collection_ratio': (total_payments / total_sales * 100) if total_sales > 0 else 0
        }
    except Exception as e:
        print(f"Error in get_sales_data: {e}")
        return {
            'total_sales': 0,
            'regular_sales': 0,
            'challan_sales': 0,
            'invoice_count': 0,
            'avg_invoice_value': 0,
            'payments_received': 0,
            'collection_ratio': 0
        }

def get_purchase_data(start_date, end_date):
    """Get comprehensive purchase data"""
    
    try:
        # Regular Purchase Invoices
        regular_purchases_total = PurchaseMaster.objects.filter(
            purchase_entry_date__date__range=[start_date, end_date]
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        regular_purchases_count = PurchaseMaster.objects.filter(
            purchase_entry_date__date__range=[start_date, end_date]
        ).count()
        
        # Payments Made
        payments_made = InvoicePaid.objects.filter(
            payment_date__range=[start_date, end_date]
        ).aggregate(total=Sum('payment_amount'))['total'] or 0
        
        total_purchases = float(regular_purchases_total)
        total_payments = float(payments_made)
        
        return {
            'total_purchases': total_purchases,
            'regular_purchases': total_purchases,
            'challan_purchases': 0,
            'invoice_count': regular_purchases_count,
            'avg_invoice_value': total_purchases / max(1, regular_purchases_count),
            'payments_made': total_payments,
            'payment_ratio': (total_payments / total_purchases * 100) if total_purchases > 0 else 0
        }
    except Exception as e:
        print(f"Error in get_purchase_data: {e}")
        return {
            'total_purchases': 0,
            'regular_purchases': 0,
            'challan_purchases': 0,
            'invoice_count': 0,
            'avg_invoice_value': 0,
            'payments_made': 0,
            'payment_ratio': 0
        }

def get_returns_data(start_date, end_date):
    """Get returns data"""
    
    try:
        sales_returns_total = ReturnSalesMaster.objects.filter(
            return_sale_entry_date__date__range=[start_date, end_date]
        ).aggregate(total=Sum('return_sale_total_amount'))['total'] or 0
        
        sales_returns_count = ReturnSalesMaster.objects.filter(
            return_sale_entry_date__date__range=[start_date, end_date]
        ).count()
        
        purchase_returns_total = ReturnPurchaseMaster.objects.filter(
            returnpurchase_entry_date__range=[start_date, end_date]
        ).aggregate(total=Sum('returntotal_amount'))['total'] or 0
        
        purchase_returns_count = ReturnPurchaseMaster.objects.filter(
            returnpurchase_entry_date__range=[start_date, end_date]
        ).count()
        
        return {
            'sales_returns': float(sales_returns_total),
            'purchase_returns': float(purchase_returns_total),
            'sales_return_count': sales_returns_count,
            'purchase_return_count': purchase_returns_count
        }
    except Exception as e:
        print(f"Error in get_returns_data: {e}")
        return {
            'sales_returns': 0,
            'purchase_returns': 0,
            'sales_return_count': 0,
            'purchase_return_count': 0
        }

def get_outstanding_amounts():
    """Get outstanding receivables and payables"""
    
    try:
        receivables = []
        payables = []
        
        # Simple calculation for outstanding amounts
        for invoice in SalesInvoiceMaster.objects.all()[:10]:
            try:
                total_sales = SalesMaster.objects.filter(
                    sales_invoice_no=invoice
                ).aggregate(total=Sum('sale_total_amount'))['total'] or 0
                
                balance = float(total_sales) - float(invoice.sales_invoice_paid)
                if balance > 0:
                    receivables.append({
                        'customer_name': invoice.customerid.customer_name,
                        'outstanding': balance,
                        'invoice_no': invoice.sales_invoice_no,
                        'invoice_date': invoice.sales_invoice_date
                    })
            except:
                continue
        
        for invoice in InvoiceMaster.objects.all()[:10]:
            try:
                balance = float(invoice.invoice_total) - float(invoice.invoice_paid)
                if balance > 0:
                    payables.append({
                        'supplier_name': invoice.supplierid.supplier_name,
                        'outstanding': balance,
                        'invoice_no': invoice.invoice_no,
                        'invoice_date': invoice.invoice_date
                    })
            except:
                continue
        
        # Sort and limit
        receivables.sort(key=lambda x: x['outstanding'], reverse=True)
        payables.sort(key=lambda x: x['outstanding'], reverse=True)
        receivables = receivables[:10]
        payables = payables[:10]
        
        total_receivables = sum(item['outstanding'] for item in receivables)
        total_payables = sum(item['outstanding'] for item in payables)
        
        return {
            'receivables': receivables,
            'payables': payables,
            'total_receivables': total_receivables,
            'total_payables': total_payables,
            'net_position': total_receivables - total_payables
        }
    except Exception as e:
        print(f"Error in get_outstanding_amounts: {e}")
        return {
            'receivables': [],
            'payables': [],
            'total_receivables': 0,
            'total_payables': 0,
            'net_position': 0
        }

def get_monthly_trends():
    """Get monthly sales and purchase trends for the last 12 months"""
    
    try:
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=365)
        
        # Monthly Sales
        monthly_sales = SalesMaster.objects.filter(
            sale_entry_date__date__range=[start_date, end_date]
        ).annotate(
            month=TruncMonth('sale_entry_date')
        ).values('month').annotate(
            total=Sum('sale_total_amount')
        ).order_by('month')
        
        # Monthly Purchases
        monthly_purchases = PurchaseMaster.objects.filter(
            purchase_entry_date__date__range=[start_date, end_date]
        ).annotate(
            month=TruncMonth('purchase_entry_date')
        ).values('month').annotate(
            total=Sum('total_amount')
        ).order_by('month')
        
        return {
            'sales': list(monthly_sales),
            'purchases': list(monthly_purchases)
        }
    except Exception as e:
        print(f"Error in get_monthly_trends: {e}")
        return {
            'sales': [],
            'purchases': []
        }

def get_inventory_metrics():
    """Get inventory-related metrics"""
    
    try:
        # Simple inventory metrics
        total_items = ProductMaster.objects.count()
        
        # Fast moving items (top 10 by sales volume)
        fast_moving = SalesMaster.objects.filter(
            sale_entry_date__date__gte=timezone.now().date() - timedelta(days=30)
        ).values('productid__product_name').annotate(
            total_qty=Sum('sale_quantity')
        ).order_by('-total_qty')[:10]
        
        return {
            'total_stock_value': 0,
            'total_stock_qty': 0,
            'low_stock_items': 0,
            'expired_items': 0,
            'total_items': total_items,
            'fast_moving_items': list(fast_moving)
        }
    except Exception as e:
        print(f"Error in get_inventory_metrics: {e}")
        return {
            'total_stock_value': 0,
            'total_stock_qty': 0,
            'low_stock_items': 0,
            'expired_items': 0,
            'total_items': 0,
            'fast_moving_items': []
        }

def get_profit_analysis(start_date, end_date):
    """Get detailed profit analysis"""
    
    try:
        # Product-wise profit analysis
        product_profits = SalesMaster.objects.filter(
            sale_entry_date__date__range=[start_date, end_date]
        ).values('productid__product_name').annotate(
            sales_amount=Sum('sale_total_amount'),
            sales_qty=Sum('sale_quantity')
        ).order_by('-sales_amount')[:10]
        
        # Customer-wise profit analysis
        customer_profits = SalesMaster.objects.filter(
            sale_entry_date__date__range=[start_date, end_date]
        ).values('customerid__customer_name').annotate(
            sales_amount=Sum('sale_total_amount'),
            invoice_count=Count('sales_invoice_no', distinct=True)
        ).order_by('-sales_amount')[:10]
        
        return {
            'top_products': list(product_profits),
            'top_customers': list(customer_profits)
        }
    except Exception as e:
        print(f"Error in get_profit_analysis: {e}")
        return {
            'top_products': [],
            'top_customers': []
        }

def get_cash_flow_data(start_date, end_date):
    """Get cash flow data"""
    
    try:
        # Daily cash flow for the period
        daily_inflow = SalesInvoicePaid.objects.filter(
            sales_payment_date__range=[start_date, end_date]
        ).annotate(
            date=TruncDate('sales_payment_date')
        ).values('date').annotate(
            amount=Sum('sales_payment_amount')
        ).order_by('date')
        
        daily_outflow = InvoicePaid.objects.filter(
            payment_date__range=[start_date, end_date]
        ).annotate(
            date=TruncDate('payment_date')
        ).values('date').annotate(
            amount=Sum('payment_amount')
        ).order_by('date')
        
        return {
            'daily_inflow': list(daily_inflow),
            'daily_outflow': list(daily_outflow)
        }
    except Exception as e:
        print(f"Error in get_cash_flow_data: {e}")
        return {
            'daily_inflow': [],
            'daily_outflow': []
        }

def get_top_performers(start_date, end_date):
    """Get top performing products, customers, and suppliers"""
    
    try:
        # Top products by sales value
        top_products = SalesMaster.objects.filter(
            sale_entry_date__date__range=[start_date, end_date]
        ).values('productid__product_name', 'productid__product_company').annotate(
            total_sales=Sum('sale_total_amount'),
            total_qty=Sum('sale_quantity')
        ).order_by('-total_sales')[:10]
        
        # Top customers by purchase value
        top_customers = SalesMaster.objects.filter(
            sale_entry_date__date__range=[start_date, end_date]
        ).values('customerid__customer_name').annotate(
            total_purchase=Sum('sale_total_amount'),
            invoice_count=Count('sales_invoice_no', distinct=True)
        ).order_by('-total_purchase')[:10]
        
        # Top suppliers by purchase value
        top_suppliers = PurchaseMaster.objects.filter(
            purchase_entry_date__date__range=[start_date, end_date]
        ).values('product_supplierid__supplier_name').annotate(
            total_purchase=Sum('total_amount'),
            invoice_count=Count('product_invoiceid', distinct=True)
        ).order_by('-total_purchase')[:10]
        
        return {
            'products': list(top_products),
            'customers': list(top_customers),
            'suppliers': list(top_suppliers)
        }
    except Exception as e:
        print(f"Error in get_top_performers: {e}")
        return {
            'products': [],
            'customers': [],
            'suppliers': []
        }

@login_required
def export_financial_pdf(request):
    """Export financial report as PDF"""
    
    # Get the same data as the main report
    end_date = request.GET.get('end_date')
    start_date = request.GET.get('start_date')
    
    if not end_date:
        end_date = timezone.now().date()
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    if not start_date:
        start_date = end_date - timedelta(days=30)
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    
    # Get all data
    sales_data = get_sales_data(start_date, end_date)
    purchase_data = get_purchase_data(start_date, end_date)
    returns_data = get_returns_data(start_date, end_date)
    outstanding_data = get_outstanding_amounts()
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1
    )
    story.append(Paragraph(f"Financial Report ({start_date} to {end_date})", title_style))
    story.append(Spacer(1, 20))
    
    # Summary Table
    summary_data = [
        ['Metric', 'Amount (₹)'],
        ['Gross Sales', f"{sales_data['total_sales']:,.2f}"],
        ['Gross Purchases', f"{purchase_data['total_purchases']:,.2f}"],
        ['Sales Returns', f"{returns_data['sales_returns']:,.2f}"],
        ['Purchase Returns', f"{returns_data['purchase_returns']:,.2f}"],
        ['Net Sales', f"{sales_data['total_sales'] - returns_data['sales_returns']:,.2f}"],
        ['Net Purchases', f"{purchase_data['total_purchases'] - returns_data['purchase_returns']:,.2f}"],
        ['Gross Profit', f"{(sales_data['total_sales'] - returns_data['sales_returns']) - (purchase_data['total_purchases'] - returns_data['purchase_returns']):,.2f}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 30))
    
    # Outstanding Amounts
    story.append(Paragraph("Outstanding Receivables (Top 5)", styles['Heading2']))
    receivables_data = [['Customer', 'Amount (₹)']]
    for item in outstanding_data['receivables'][:5]:
        receivables_data.append([item['customer_name'], f"{item['outstanding']:,.2f}"])
    
    receivables_table = Table(receivables_data, colWidths=[3*inch, 2*inch])
    receivables_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(receivables_table)
    story.append(Spacer(1, 20))
    
    story.append(Paragraph("Outstanding Payables (Top 5)", styles['Heading2']))
    payables_data = [['Supplier', 'Amount (₹)']]
    for item in outstanding_data['payables'][:5]:
        payables_data.append([item['supplier_name'], f"{item['outstanding']:,.2f}"])
    
    payables_table = Table(payables_data, colWidths=[3*inch, 2*inch])
    payables_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(payables_table)
    
    doc.build(story)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="financial_report_{start_date}_to_{end_date}.pdf"'
    
    return response

@login_required
def export_financial_excel(request):
    """Export financial report as Excel"""
    
    # Get the same data as the main report
    end_date = request.GET.get('end_date')
    start_date = request.GET.get('start_date')
    
    if not end_date:
        end_date = timezone.now().date()
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    if not start_date:
        start_date = end_date - timedelta(days=30)
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    
    # Get all data
    sales_data = get_sales_data(start_date, end_date)
    purchase_data = get_purchase_data(start_date, end_date)
    returns_data = get_returns_data(start_date, end_date)
    outstanding_data = get_outstanding_amounts()
    monthly_trends = get_monthly_trends()
    inventory_metrics = get_inventory_metrics()
    top_performers = get_top_performers(start_date, end_date)
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Financial Summary"
    
    # Headers
    ws_summary['A1'] = f"Financial Report ({start_date} to {end_date})"
    ws_summary['A1'].font = Font(size=16, bold=True)
    ws_summary.merge_cells('A1:B1')
    
    # Summary data
    summary_data = [
        ['Metric', 'Amount (₹)'],
        ['Gross Sales', sales_data['total_sales']],
        ['Gross Purchases', purchase_data['total_purchases']],
        ['Sales Returns', returns_data['sales_returns']],
        ['Purchase Returns', returns_data['purchase_returns']],
        ['Net Sales', sales_data['total_sales'] - returns_data['sales_returns']],
        ['Net Purchases', purchase_data['total_purchases'] - returns_data['purchase_returns']],
        ['Gross Profit', (sales_data['total_sales'] - returns_data['sales_returns']) - (purchase_data['total_purchases'] - returns_data['purchase_returns'])]
    ]
    
    for row_num, row_data in enumerate(summary_data, 3):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_num, column=col_num, value=value)
            if row_num == 3:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Outstanding Receivables Sheet
    ws_receivables = wb.create_sheet("Outstanding Receivables")
    ws_receivables['A1'] = "Outstanding Receivables"
    ws_receivables['A1'].font = Font(size=14, bold=True)
    
    receivables_headers = ['Customer', 'Invoice No', 'Invoice Date', 'Outstanding Amount']
    for col_num, header in enumerate(receivables_headers, 1):
        cell = ws_receivables.cell(row=2, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for row_num, item in enumerate(outstanding_data['receivables'], 3):
        ws_receivables.cell(row=row_num, column=1, value=item['customer_name'])
        ws_receivables.cell(row=row_num, column=2, value=item['invoice_no'])
        ws_receivables.cell(row=row_num, column=3, value=item['invoice_date'])
        ws_receivables.cell(row=row_num, column=4, value=float(item['outstanding']))
    
    # Outstanding Payables Sheet
    ws_payables = wb.create_sheet("Outstanding Payables")
    ws_payables['A1'] = "Outstanding Payables"
    ws_payables['A1'].font = Font(size=14, bold=True)
    
    payables_headers = ['Supplier', 'Invoice No', 'Invoice Date', 'Outstanding Amount']
    for col_num, header in enumerate(payables_headers, 1):
        cell = ws_payables.cell(row=2, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for row_num, item in enumerate(outstanding_data['payables'], 3):
        ws_payables.cell(row=row_num, column=1, value=item['supplier_name'])
        ws_payables.cell(row=row_num, column=2, value=item['invoice_no'])
        ws_payables.cell(row=row_num, column=3, value=item['invoice_date'])
        ws_payables.cell(row=row_num, column=4, value=float(item['outstanding']))
    
    # Top Performers Sheet
    ws_performers = wb.create_sheet("Top Performers")
    ws_performers['A1'] = "Top Performing Products"
    ws_performers['A1'].font = Font(size=14, bold=True)
    
    # Top Products
    product_headers = ['Product Name', 'Company', 'Sales Amount', 'Quantity Sold']
    for col_num, header in enumerate(product_headers, 1):
        cell = ws_performers.cell(row=2, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for row_num, item in enumerate(top_performers['products'], 3):
        ws_performers.cell(row=row_num, column=1, value=item['productid__product_name'])
        ws_performers.cell(row=row_num, column=2, value=item['productid__product_company'])
        ws_performers.cell(row=row_num, column=3, value=float(item['total_sales']))
        ws_performers.cell(row=row_num, column=4, value=float(item['total_qty']))
    
    # Top Customers
    start_row = len(top_performers['products']) + 5
    ws_performers.cell(row=start_row, column=1, value="Top Customers")
    ws_performers.cell(row=start_row, column=1).font = Font(size=14, bold=True)
    
    customer_headers = ['Customer Name', 'Total Purchase', 'Invoice Count']
    for col_num, header in enumerate(customer_headers, 1):
        cell = ws_performers.cell(row=start_row + 1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for row_num, item in enumerate(top_performers['customers'], start_row + 2):
        ws_performers.cell(row=row_num, column=1, value=item['customerid__customer_name'])
        ws_performers.cell(row=row_num, column=2, value=float(item['total_purchase']))
        ws_performers.cell(row=row_num, column=3, value=item['invoice_count'])
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="financial_report_{start_date}_to_{end_date}.xlsx"'
    
    wb.save(response)
    return response

@login_required
def financial_dashboard_api(request):
    """API endpoint for financial dashboard data"""
    
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)
    
    # Get key metrics
    sales_data = get_sales_data(start_date, end_date)
    purchase_data = get_purchase_data(start_date, end_date)
    returns_data = get_returns_data(start_date, end_date)
    outstanding_data = get_outstanding_amounts()
    
    data = {
        'sales': float(sales_data['total_sales']),
        'purchases': float(purchase_data['total_purchases']),
        'gross_profit': float(sales_data['total_sales'] - purchase_data['total_purchases']),
        'total_receivables': float(outstanding_data['total_receivables']),
        'total_payables': float(outstanding_data['total_payables']),
        'collection_ratio': sales_data['collection_ratio'],
        'payment_ratio': purchase_data['payment_ratio']
    }
    
    return JsonResponse(data)