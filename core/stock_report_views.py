from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Q
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from datetime import datetime, timedelta
import csv
import json
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import ProductMaster
from .stock_manager import StockManager


@login_required
def stock_statement_report(request):
    """
    Comprehensive stock statement report showing opening, received, sold, and balance
    """
    # Get filter parameters
    search_query = request.GET.get('search', '').strip()
    category_filter = request.GET.get('category', '')
    company_filter = request.GET.get('company', '')
    stock_status = request.GET.get('stock_status', '')  # all, in_stock, low_stock, out_of_stock
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Base query
    products_query = ProductMaster.objects.all().order_by('product_name')
    
    # Check if filters applied
    has_filters = any([search_query, category_filter, company_filter, stock_status and stock_status != 'all', date_from, date_to])
    
    # Apply filters
    if search_query:
        products_query = products_query.filter(
            Q(product_name__icontains=search_query) |
            Q(product_company__icontains=search_query) |
            Q(product_salt__icontains=search_query) |
            Q(product_barcode__icontains=search_query)
        )
    
    if category_filter:
        products_query = products_query.filter(product_category__icontains=category_filter)
    
    if company_filter:
        products_query = products_query.filter(product_company__icontains=company_filter)
    
    # Limit to 100 products if no filters
    if not has_filters:
        products_query = products_query[:100]
    
    # Get unique categories and companies for filter dropdowns
    categories = ProductMaster.objects.values_list('product_category', flat=True).distinct().order_by('product_category')
    companies = ProductMaster.objects.values_list('product_company', flat=True).distinct().order_by('product_company')
    
    # Pagination
    paginator = Paginator(products_query, 25)  # 25 products per page
    page_number = request.GET.get('page')
    products_page = paginator.get_page(page_number)
    
    # OPTIMIZED: Bulk fetch all data in 4 queries instead of N+1
    from .models import PurchaseMaster, SalesMaster, ReturnPurchaseMaster, ReturnSalesMaster
    
    product_ids = [p.productid for p in products_page]
    purchases = PurchaseMaster.objects.filter(productid__in=product_ids).values('productid').annotate(qty=Sum('product_quantity'), avg_mrp=Sum('product_MRP')/Sum('product_quantity'))
    sales = SalesMaster.objects.filter(productid__in=product_ids).values('productid').annotate(qty=Sum('sale_quantity'))
    purchase_returns = ReturnPurchaseMaster.objects.filter(returnproductid__in=product_ids).values('returnproductid').annotate(qty=Sum('returnproduct_quantity'))
    sales_returns = ReturnSalesMaster.objects.filter(return_productid__in=product_ids).values('return_productid').annotate(qty=Sum('return_sale_quantity'))
    
    purchase_dict = {p['productid']: {'qty': p['qty'] or 0, 'mrp': p['avg_mrp'] or 0} for p in purchases}
    sales_dict = {s['productid']: s['qty'] or 0 for s in sales}
    pr_dict = {pr['returnproductid']: pr['qty'] or 0 for pr in purchase_returns}
    sr_dict = {sr['return_productid']: sr['qty'] or 0 for sr in sales_returns}
    
    stock_data = []
    total_opening = 0
    total_received = 0
    total_sold = 0
    total_balance = 0
    total_value = 0
    
    for product in products_page:
        try:
            # OPTIMIZED: Use pre-fetched data
            pid = product.productid
            
            # Calculate opening stock (stock before start date)
            if date_from:
                try:
                    from datetime import datetime
                    start_date = datetime.strptime(date_from, '%Y-%m-%d')
                    
                    # Get transactions before start date
                    from .models import PurchaseMaster, SalesMaster, ReturnPurchaseMaster, ReturnSalesMaster
                    
                    prev_purchases = PurchaseMaster.objects.filter(
                        productid=pid,
                        purchase_entry_date__lt=start_date
                    ).aggregate(qty=Sum('product_quantity'))['qty'] or 0
                    
                    prev_sales = SalesMaster.objects.filter(
                        productid=pid,
                        sale_entry_date__lt=start_date
                    ).aggregate(qty=Sum('sale_quantity'))['qty'] or 0
                    
                    prev_pr = ReturnPurchaseMaster.objects.filter(
                        returnproductid=pid,
                        returnpurchase_entry_date__lt=start_date
                    ).aggregate(qty=Sum('returnproduct_quantity'))['qty'] or 0
                    
                    prev_sr = ReturnSalesMaster.objects.filter(
                        return_productid=pid,
                        return_entry_date__lt=start_date
                    ).aggregate(qty=Sum('return_sale_quantity'))['qty'] or 0
                    
                    opening_stock = (prev_purchases + prev_sr) - (prev_sales + prev_pr)
                except:
                    opening_stock = 0
            else:
                opening_stock = 0
            
            received_stock = purchase_dict.get(pid, {}).get('qty', 0) + sr_dict.get(pid, 0)
            sold_stock = sales_dict.get(pid, 0) + pr_dict.get(pid, 0)
            balance_stock = opening_stock + received_stock - sold_stock
            avg_mrp = purchase_dict.get(pid, {}).get('mrp', 0)
            stock_value = balance_stock * avg_mrp
            
            # Determine stock status
            if balance_stock <= 0:
                status = 'out_of_stock'
                status_label = 'Out of Stock'
                status_class = 'danger'
            elif balance_stock < 10:
                status = 'low_stock'
                status_label = 'Low Stock'
                status_class = 'warning'
            else:
                status = 'in_stock'
                status_label = 'In Stock'
                status_class = 'success'
            
            # Apply stock status filter
            if stock_status and stock_status != 'all' and stock_status != status:
                continue
            
            stock_item = {
                'product': product,
                'opening_stock': opening_stock,
                'received_stock': received_stock,
                'sold_stock': sold_stock,
                'balance_stock': balance_stock,
                'avg_mrp': avg_mrp,
                'stock_value': stock_value,
                'status': status,
                'status_label': status_label,
                'status_class': status_class,
                'batches': []  # OPTIMIZED: Skip batch details for performance
            }
            
            stock_data.append(stock_item)
            
            # Add to totals
            total_opening += opening_stock
            total_received += received_stock
            total_sold += sold_stock
            total_balance += balance_stock
            total_value += stock_value
            
        except Exception as e:
            print(f"Error processing stock for {product.product_name}: {e}")
            # Add error entry
            stock_data.append({
                'product': product,
                'opening_stock': 0,
                'received_stock': 0,
                'sold_stock': 0,
                'balance_stock': 0,
                'avg_mrp': 0,
                'stock_value': 0,
                'status': 'error',
                'status_label': 'Error',
                'status_class': 'secondary',
                'batches': []
            })
    
    # Handle export requests
    export_type = request.GET.get('export')
    if export_type == 'pdf':
        return export_stock_statement_pdf(request)
    elif export_type == 'excel':
        return export_stock_statement_excel(request)
    
    context = {
        'stock_data': stock_data,
        'products_page': products_page,
        'search_query': search_query,
        'category_filter': category_filter,
        'company_filter': company_filter,
        'stock_status': stock_status,
        'date_from': date_from,
        'date_to': date_to,
        'categories': [cat for cat in categories if cat],
        'companies': [comp for comp in companies if comp],
        'totals': {
            'opening': total_opening,
            'received': total_received,
            'sold': total_sold,
            'balance': total_balance,
            'value': total_value
        },
        'title': 'Stock Statement Report'
    }
    
    return render(request, 'reports/stock_statement_report.html', context)


@login_required
def export_stock_statement_pdf(request):
    """Export stock statement to PDF"""
    # Get filter parameters (same as main view)
    search_query = request.GET.get('search', '').strip()
    category_filter = request.GET.get('category', '')
    company_filter = request.GET.get('company', '')
    stock_status = request.GET.get('stock_status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    page_number = request.GET.get('page', '1')
    
    # Base query
    products_query = ProductMaster.objects.all().order_by('product_name')
    
    # Check if any filter is applied
    has_filters = any([search_query, category_filter, company_filter, stock_status and stock_status != 'all', date_from, date_to])
    
    # Apply filters
    if search_query:
        products_query = products_query.filter(
            Q(product_name__icontains=search_query) |
            Q(product_company__icontains=search_query) |
            Q(product_salt__icontains=search_query) |
            Q(product_barcode__icontains=search_query)
        )
    
    if category_filter:
        products_query = products_query.filter(product_category__icontains=category_filter)
    
    if company_filter:
        products_query = products_query.filter(product_company__icontains=company_filter)
    
    # If no filters applied, use pagination (only current page)
    if not has_filters:
        paginator = Paginator(products_query, 25)
        try:
            products_query = paginator.page(page_number)
        except:
            products_query = paginator.page(1)
    
    # Process stock data
    stock_data = []
    totals = {'opening': 0, 'received': 0, 'sold': 0, 'balance': 0, 'value': 0}
    
    # OPTIMIZED: Bulk queries for PDF export
    products_list = products_query.object_list if hasattr(products_query, 'object_list') else products_query
    product_ids = [p.productid for p in products_list]
    
    from .models import PurchaseMaster, SalesMaster, ReturnPurchaseMaster, ReturnSalesMaster
    purchases = PurchaseMaster.objects.filter(productid__in=product_ids).values('productid').annotate(qty=Sum('product_quantity'), avg_mrp=Sum('product_MRP')/Sum('product_quantity'))
    sales = SalesMaster.objects.filter(productid__in=product_ids).values('productid').annotate(qty=Sum('sale_quantity'))
    purchase_returns = ReturnPurchaseMaster.objects.filter(returnproductid__in=product_ids).values('returnproductid').annotate(qty=Sum('returnproduct_quantity'))
    sales_returns = ReturnSalesMaster.objects.filter(return_productid__in=product_ids).values('return_productid').annotate(qty=Sum('return_sale_quantity'))
    
    purchase_dict = {p['productid']: {'qty': p['qty'] or 0, 'mrp': p['avg_mrp'] or 0} for p in purchases}
    sales_dict = {s['productid']: s['qty'] or 0 for s in sales}
    pr_dict = {pr['returnproductid']: pr['qty'] or 0 for pr in purchase_returns}
    sr_dict = {sr['return_productid']: sr['qty'] or 0 for sr in sales_returns}
    
    for product in products_list:
        try:
            pid = product.productid
            opening_stock = 0
            received_stock = purchase_dict.get(pid, {}).get('qty', 0) + sr_dict.get(pid, 0)
            sold_stock = sales_dict.get(pid, 0) + pr_dict.get(pid, 0)
            balance_stock = received_stock - sold_stock
            avg_mrp = purchase_dict.get(pid, {}).get('mrp', 0)
            stock_value = balance_stock * avg_mrp
            
            if balance_stock <= 0:
                status_label = 'Out of Stock'
            elif balance_stock < 10:
                status_label = 'Low Stock'
            else:
                status_label = 'In Stock'
            
            # Apply stock status filter
            if stock_status and stock_status != 'all':
                if balance_stock <= 0 and stock_status != 'out_of_stock':
                    continue
                elif balance_stock > 0 and balance_stock < 10 and stock_status != 'low_stock':
                    continue
                elif balance_stock >= 10 and stock_status != 'in_stock':
                    continue
            
            stock_data.append({
                'product': product,
                'opening_stock': opening_stock,
                'received_stock': received_stock,
                'sold_stock': sold_stock,
                'balance_stock': balance_stock,
                'avg_mrp': avg_mrp,
                'stock_value': stock_value,
                'status_label': status_label
            })
            
            totals['opening'] += opening_stock
            totals['received'] += received_stock
            totals['sold'] += sold_stock
            totals['balance'] += balance_stock
            totals['value'] += stock_value
        except:
            continue
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="stock_statement_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=20)
    elements = []
    styles = getSampleStyleSheet()
    
    # Get pharmacy details from database
    from .models import Pharmacy_Details
    try:
        pharmacy = Pharmacy_Details.objects.first()
    except:
        pharmacy = None
    
    # Add pharmacy details header
    if pharmacy:
        pharma_name = Paragraph(f"<b>{pharmacy.pharmaname}</b>", ParagraphStyle('PharmaName', parent=styles['Heading1'], fontSize=16, alignment=1))
        elements.append(pharma_name)
        elements.append(Spacer(1, 6))
        
        if pharmacy.proprietorname:
            proprietor = Paragraph(f"Proprietor: {pharmacy.proprietorname}", ParagraphStyle('Info', parent=styles['Normal'], fontSize=10, alignment=1))
            elements.append(proprietor)
            elements.append(Spacer(1, 4))
        
        contact_info = []
        if pharmacy.proprietorcontact:
            contact_info.append(f"Contact: {pharmacy.proprietorcontact}")
        if pharmacy.proprietoremail:
            contact_info.append(f"Email: {pharmacy.proprietoremail}")
        
        if contact_info:
            contact = Paragraph(" | ".join(contact_info), ParagraphStyle('Contact', parent=styles['Normal'], fontSize=9, alignment=1))
            elements.append(contact)
            elements.append(Spacer(1, 4))
        
        if pharmacy.pharmaweburl:
            website = Paragraph(f"Website: {pharmacy.pharmaweburl}", ParagraphStyle('Website', parent=styles['Normal'], fontSize=9, alignment=1))
            elements.append(website)
            elements.append(Spacer(1, 10))
    
    # Report title
    title = Paragraph("<b>Stock Statement Report</b>", ParagraphStyle('Title', parent=styles['Heading2'], fontSize=14, alignment=1))
    elements.append(title)
    elements.append(Spacer(1, 8))
    
    date_text = f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    elements.append(Paragraph(date_text, ParagraphStyle('Date', parent=styles['Normal'], fontSize=9, alignment=1)))
    elements.append(Spacer(1, 12))
    
    # Summary
    summary_data = [[
        f"Opening: {int(totals['opening'])}",
        f"Received: {int(totals['received'])}",
        f"Sold: {int(totals['sold'])}",
        f"Balance: {int(totals['balance'])}",
        f"Value: ₹{totals['value']:.2f}"
    ]]
    summary_table = Table(summary_data, colWidths=[1.8*inch]*5)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 15))
    
    # Main table (removed MRP and Status columns)
    data = [['Product', 'Company', 'Opening', 'Received', 'Sold', 'Balance', 'Value']]
    for item in stock_data:
        data.append([
            item['product'].product_name[:35],
            item['product'].product_company[:20],
            str(int(item['opening_stock'])),
            str(int(item['received_stock'])),
            str(int(item['sold_stock'])),
            str(int(item['balance_stock'])),
            f"₹{item['stock_value']:.2f}"
        ])
    
    table = Table(data, colWidths=[3*inch, 1.8*inch, 0.9*inch, 0.9*inch, 0.9*inch, 0.9*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    elements.append(table)
    
    doc.build(elements)
    return response


def export_stock_statement_excel(request, stock_data=None):
    """Export stock statement to Excel"""
    # If stock_data not provided, fetch with filters
    if stock_data is None:
        search_query = request.GET.get('search', '').strip()
        category_filter = request.GET.get('category', '')
        company_filter = request.GET.get('company', '')
        stock_status = request.GET.get('stock_status', '')
        page_number = request.GET.get('page', '1')
        
        products_query = ProductMaster.objects.all().order_by('product_name')
        
        # Check if any filter is applied
        has_filters = any([search_query, category_filter, company_filter, stock_status and stock_status != 'all'])
        
        if search_query:
            products_query = products_query.filter(
                Q(product_name__icontains=search_query) |
                Q(product_company__icontains=search_query) |
                Q(product_salt__icontains=search_query) |
                Q(product_barcode__icontains=search_query)
            )
        
        if category_filter:
            products_query = products_query.filter(product_category__icontains=category_filter)
        
        if company_filter:
            products_query = products_query.filter(product_company__icontains=company_filter)
        
        # If no filters applied, use pagination (only current page)
        if not has_filters:
            paginator = Paginator(products_query, 25)
            try:
                products_query = paginator.page(page_number)
            except:
                products_query = paginator.page(1)
        
        # OPTIMIZED: Bulk queries for Excel export
        products_list = products_query.object_list if hasattr(products_query, 'object_list') else products_query
        product_ids = [p.productid for p in products_list]
        
        from .models import PurchaseMaster, SalesMaster, ReturnPurchaseMaster, ReturnSalesMaster
        purchases = PurchaseMaster.objects.filter(productid__in=product_ids).values('productid').annotate(qty=Sum('product_quantity'), avg_mrp=Sum('product_MRP')/Sum('product_quantity'))
        sales = SalesMaster.objects.filter(productid__in=product_ids).values('productid').annotate(qty=Sum('sale_quantity'))
        purchase_returns = ReturnPurchaseMaster.objects.filter(returnproductid__in=product_ids).values('returnproductid').annotate(qty=Sum('returnproduct_quantity'))
        sales_returns = ReturnSalesMaster.objects.filter(return_productid__in=product_ids).values('return_productid').annotate(qty=Sum('return_sale_quantity'))
        
        purchase_dict = {p['productid']: {'qty': p['qty'] or 0, 'mrp': p['avg_mrp'] or 0} for p in purchases}
        sales_dict = {s['productid']: s['qty'] or 0 for s in sales}
        pr_dict = {pr['returnproductid']: pr['qty'] or 0 for pr in purchase_returns}
        sr_dict = {sr['return_productid']: sr['qty'] or 0 for sr in sales_returns}
        
        stock_data = []
        for product in products_list:
            try:
                pid = product.productid
                opening_stock = 0
                received_stock = purchase_dict.get(pid, {}).get('qty', 0) + sr_dict.get(pid, 0)
                sold_stock = sales_dict.get(pid, 0) + pr_dict.get(pid, 0)
                balance_stock = received_stock - sold_stock
                avg_mrp = purchase_dict.get(pid, {}).get('mrp', 0)
                stock_value = balance_stock * avg_mrp
                
                if balance_stock <= 0:
                    status_label = 'Out of Stock'
                elif balance_stock < 10:
                    status_label = 'Low Stock'
                else:
                    status_label = 'In Stock'
                
                # Apply stock status filter
                if stock_status and stock_status != 'all':
                    if balance_stock <= 0 and stock_status != 'out_of_stock':
                        continue
                    elif balance_stock > 0 and balance_stock < 10 and stock_status != 'low_stock':
                        continue
                    elif balance_stock >= 10 and stock_status != 'in_stock':
                        continue
                
                stock_data.append({
                    'product': product,
                    'opening_stock': opening_stock,
                    'received_stock': received_stock,
                    'sold_stock': sold_stock,
                    'balance_stock': balance_stock,
                    'avg_mrp': avg_mrp,
                    'stock_value': stock_value,
                    'status_label': status_label
                })
            except:
                continue
    
    # Original Excel export code

    """Export stock statement to Excel"""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'stock_statement_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Stock Statement'
    
    # Get pharmacy details and date range
    from .models import Pharmacy_Details
    try:
        pharmacy = Pharmacy_Details.objects.first()
    except:
        pharmacy = None
    
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Define styles
    title_font = Font(bold=True, size=16, color='000080')
    info_font = Font(size=10)
    subtitle_font = Font(bold=True, size=12)
    date_font = Font(size=10, italic=True, bold=True)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    center_alignment = Alignment(horizontal='center')
    
    current_row = 1
    
    # Add pharmacy details
    if pharmacy:
        worksheet.cell(row=current_row, column=1, value=pharmacy.pharmaname).font = title_font
        worksheet.cell(row=current_row, column=1).alignment = center_alignment
        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        current_row += 1
        
        if pharmacy.proprietorname:
            worksheet.cell(row=current_row, column=1, value=f"Proprietor: {pharmacy.proprietorname}").font = info_font
            worksheet.cell(row=current_row, column=1).alignment = center_alignment
            worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            current_row += 1
        
        if pharmacy.proprietorcontact:
            worksheet.cell(row=current_row, column=1, value=f"Contact: {pharmacy.proprietorcontact}").font = info_font
            worksheet.cell(row=current_row, column=1).alignment = center_alignment
            worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            current_row += 1
        
        if pharmacy.proprietoremail:
            worksheet.cell(row=current_row, column=1, value=f"Email: {pharmacy.proprietoremail}").font = info_font
            worksheet.cell(row=current_row, column=1).alignment = center_alignment
            worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            current_row += 1
        
        if pharmacy.pharmaweburl:
            worksheet.cell(row=current_row, column=1, value=f"Website: {pharmacy.pharmaweburl}").font = info_font
            worksheet.cell(row=current_row, column=1).alignment = center_alignment
            worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            current_row += 1
        
        current_row += 1
    
    # Add report title
    worksheet.cell(row=current_row, column=1, value="STOCK STATEMENT REPORT").font = subtitle_font
    worksheet.cell(row=current_row, column=1).alignment = center_alignment
    worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    current_row += 1
    
    # Add date range
    if date_from and date_to:
        try:
            from_date_obj = datetime.strptime(date_from, '%Y-%m-%d')
            to_date_obj = datetime.strptime(date_to, '%Y-%m-%d')
            date_range = f"Period: {from_date_obj.strftime('%d/%m/%Y')} to {to_date_obj.strftime('%d/%m/%Y')}"
        except:
            date_range = f"Period: {date_from} to {date_to}"
        worksheet.cell(row=current_row, column=1, value=date_range).font = date_font
        worksheet.cell(row=current_row, column=1).alignment = center_alignment
        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        current_row += 1
    
    # Add generated date
    generated_date = f"Generated on: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    worksheet.cell(row=current_row, column=1, value=generated_date).font = date_font
    worksheet.cell(row=current_row, column=1).alignment = center_alignment
    worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    current_row += 2
    
    # Add column headers
    headers = [
        'Product Name', 'Packing', 'Opening Stock',
        'Received', 'Sold', 'Balance', 'Stock Value'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    current_row += 1
    
    # Add data
    for item in stock_data:
        worksheet.cell(row=current_row, column=1, value=item['product'].product_name)
        worksheet.cell(row=current_row, column=2, value=item['product'].product_packing)
        worksheet.cell(row=current_row, column=3, value=item['opening_stock'])
        worksheet.cell(row=current_row, column=4, value=item['received_stock'])
        worksheet.cell(row=current_row, column=5, value=item['sold_stock'])
        worksheet.cell(row=current_row, column=6, value=item['balance_stock'])
        worksheet.cell(row=current_row, column=7, value=item['stock_value'])
        current_row += 1
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(response)
    return response


@login_required
def stock_statement_batch_detail(request, product_id):
    """Get detailed batch information for a product"""
    try:
        product = ProductMaster.objects.get(productid=product_id)
        stock_summary = StockManager.get_stock_summary(product_id)
        
        batch_details = []
        for batch in stock_summary['batches']:
            # Get additional batch information
            from .models import PurchaseMaster, SaleRateMaster
            
            # Get MRP and rates
            purchase = PurchaseMaster.objects.filter(
                productid=product_id,
                product_batch_no=batch['batch_no']
            ).first()
            
            try:
                rates = SaleRateMaster.objects.get(
                    productid=product_id,
                    product_batch_no=batch['batch_no']
                )
                rate_A = rates.rate_A
                rate_B = rates.rate_B
                rate_C = rates.rate_C
            except SaleRateMaster.DoesNotExist:
                rate_A = rate_B = rate_C = 0
            
            batch_details.append({
                'batch_no': batch['batch_no'],
                'expiry': batch['expiry'],
                'stock': batch['stock'],
                'purchased': batch['purchased'],
                'sold': batch['sold'],
                'purchase_returns': batch['purchase_returns'],
                'sales_returns': batch['sales_returns'],
                'mrp': purchase.product_MRP if purchase else 0,
                'rate_A': rate_A,
                'rate_B': rate_B,
                'rate_C': rate_C
            })
        
        return JsonResponse({
            'success': True,
            'product_name': product.product_name,
            'product_company': product.product_company,
            'batches': batch_details
        })
        
    except ProductMaster.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Product not found'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })