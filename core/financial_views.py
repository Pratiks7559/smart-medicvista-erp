from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, F, Q, FloatField, ExpressionWrapper
from django.http import HttpResponse
from datetime import datetime, timedelta
from decimal import Decimal
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .models import SalesMaster, PurchaseMaster, ProductMaster, SupplierChallanMaster, CustomerChallanMaster

@login_required
def financial_report(request):
    """Financial Report with Profit Calculation"""
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    product_id = request.GET.get('product_id')
    product_search = request.GET.get('product_search', '')
    
    # Base querysets with ordering
    sales_query = SalesMaster.objects.select_related('productid', 'sales_invoice_no', 'customerid').order_by('-sale_entry_date')
    purchase_query = PurchaseMaster.objects.select_related('productid', 'product_invoiceid', 'product_supplierid').order_by('-purchase_entry_date')
    supplier_challan_query = SupplierChallanMaster.objects.select_related('product_id', 'product_suppliername', 'product_challan_id').filter(product_challan_id__is_invoiced=False).order_by('-challan_entry_date')
    customer_challan_query = CustomerChallanMaster.objects.select_related('product_id', 'customer_name', 'customer_challan_id').order_by('-sales_entry_date')
    
    # Apply date filters
    if start_date:
        try:
            sales_query = sales_query.filter(sale_entry_date__date__gte=start_date)
            purchase_query = purchase_query.filter(purchase_entry_date__date__gte=start_date)
            supplier_challan_query = supplier_challan_query.filter(challan_entry_date__date__gte=start_date)
            customer_challan_query = customer_challan_query.filter(sales_entry_date__date__gte=start_date)
        except:
            pass
    if end_date:
        try:
            sales_query = sales_query.filter(sale_entry_date__date__lte=end_date)
            purchase_query = purchase_query.filter(purchase_entry_date__date__lte=end_date)
            supplier_challan_query = supplier_challan_query.filter(challan_entry_date__date__lte=end_date)
            customer_challan_query = customer_challan_query.filter(sales_entry_date__date__lte=end_date)
        except:
            pass
    
    # Apply product filter
    if product_id and str(product_id).strip() and str(product_id).strip() != '':
        try:
            pid = int(product_id)
            sales_query = sales_query.filter(productid_id=pid)
            purchase_query = purchase_query.filter(productid_id=pid)
            supplier_challan_query = supplier_challan_query.filter(product_id=pid)
            customer_challan_query = customer_challan_query.filter(product_id=pid)
        except:
            if product_search and product_search.strip():
                search_term = product_search.strip()
                sales_query = sales_query.filter(product_name__icontains=search_term)
                purchase_query = purchase_query.filter(product_name__icontains=search_term)
                supplier_challan_query = supplier_challan_query.filter(product_name__icontains=search_term)
                customer_challan_query = customer_challan_query.filter(product_name__icontains=search_term)
    elif product_search and product_search.strip():
        search_term = product_search.strip()
        sales_query = sales_query.filter(product_name__icontains=search_term)
        purchase_query = purchase_query.filter(product_name__icontains=search_term)
        supplier_challan_query = supplier_challan_query.filter(product_name__icontains=search_term)
        customer_challan_query = customer_challan_query.filter(product_name__icontains=search_term)
    
    # Limit data if no date filter
    if not (start_date and end_date):
        sales_query = sales_query[:500]
        purchase_query = purchase_query[:500]
        customer_challan_query = customer_challan_query[:500]
        supplier_challan_query = supplier_challan_query[:500]
    
    # Fetch data
    sales_list = list(sales_query)
    purchase_list = list(purchase_query)
    customer_challan_list = list(customer_challan_query)
    supplier_challan_list = list(supplier_challan_query)
    
    # Build batch lookup (single query)
    batch_keys = set()
    for sale in sales_list:
        batch_keys.add((sale.productid_id, sale.product_batch_no))
    for challan in customer_challan_list:
        batch_keys.add((challan.product_id_id, challan.product_batch_no))
    
    purchase_lookup = {}
    if batch_keys:
        q_objects = Q()
        for prod_id, batch_no in batch_keys:
            q_objects |= Q(productid_id=prod_id, product_batch_no=batch_no)
        purchases = PurchaseMaster.objects.filter(q_objects).values('productid_id', 'product_batch_no', 'product_purchase_rate')
        for p in purchases:
            key = (p['productid_id'], p['product_batch_no'])
            if key not in purchase_lookup:
                purchase_lookup[key] = float(p['product_purchase_rate'])
    
    # Calculate profit
    financial_data = []
    total_sales_value = 0.0
    total_purchase_cost = 0.0
    total_profit = 0.0
    total_gst = 0.0
    
    # Process Sales
    for sale in sales_list:
        purchase_rate = purchase_lookup.get((sale.productid_id, sale.product_batch_no), 0.0)
        quantity = float(sale.sale_quantity)
        sale_rate = float(sale.sale_rate)
        cgst = float(sale.sale_cgst)
        sgst = float(sale.sale_sgst)
        
        purchase_cost = purchase_rate * quantity
        sales_value = sale_rate * quantity
        gst_amount = (cgst + sgst) * quantity
        profit = (sales_value - purchase_cost) + gst_amount
        
        financial_data.append({
            'type': 'Sale',
            'date': sale.sale_entry_date,
            'invoice_no': sale.sales_invoice_no.sales_invoice_no,
            'customer': sale.customerid.customer_name,
            'product_name': sale.product_name,
            'company': sale.product_company,
            'batch_no': sale.product_batch_no,
            'quantity': quantity,
            'mrp': float(sale.product_MRP),
            'purchase_rate': purchase_rate,
            'sale_rate': sale_rate,
            'cgst': cgst,
            'sgst': sgst,
            'gst_amount': gst_amount,
            'purchase_cost': purchase_cost,
            'sales_value': sales_value,
            'profit': profit,
            'profit_percentage': (profit / sales_value * 100) if sales_value > 0 else 0
        })
        
        total_sales_value += sales_value
        total_purchase_cost += purchase_cost
        total_profit += profit
        total_gst += gst_amount
    
    # Process Customer Challans
    for challan in customer_challan_list:
        purchase_rate = purchase_lookup.get((challan.product_id_id, challan.product_batch_no), 0.0)
        quantity = float(challan.sale_quantity)
        sale_rate = float(challan.sale_rate)
        cgst = float(challan.sale_cgst)
        sgst = float(challan.sale_sgst)
        
        purchase_cost = purchase_rate * quantity
        sales_value = sale_rate * quantity
        gst_amount = (cgst + sgst) * quantity
        profit = (sales_value - purchase_cost) + gst_amount
        
        financial_data.append({
            'type': 'Customer Challan',
            'date': challan.sales_entry_date,
            'invoice_no': challan.customer_challan_no,
            'customer': challan.customer_name.customer_name,
            'product_name': challan.product_name,
            'company': challan.product_company,
            'batch_no': challan.product_batch_no,
            'quantity': quantity,
            'mrp': float(challan.product_mrp),
            'purchase_rate': purchase_rate,
            'sale_rate': sale_rate,
            'cgst': cgst,
            'sgst': sgst,
            'gst_amount': gst_amount,
            'purchase_cost': purchase_cost,
            'sales_value': sales_value,
            'profit': profit,
            'profit_percentage': (profit / sales_value * 100) if sales_value > 0 else 0
        })
        
        total_sales_value += sales_value
        total_purchase_cost += purchase_cost
        total_profit += profit
        total_gst += gst_amount
    
    # Process Purchases
    for purchase in purchase_list:
        quantity = float(purchase.product_quantity)
        purchase_rate = float(purchase.product_purchase_rate)
        cgst = float(purchase.CGST)
        sgst = float(purchase.SGST)
        
        purchase_cost = purchase_rate * quantity
        gst_amount = (cgst + sgst) * quantity
        
        financial_data.append({
            'type': 'Purchase',
            'date': purchase.purchase_entry_date,
            'invoice_no': purchase.product_invoice_no,
            'customer': purchase.product_supplierid.supplier_name,
            'product_name': purchase.product_name,
            'company': purchase.product_company,
            'batch_no': purchase.product_batch_no,
            'quantity': quantity,
            'mrp': float(purchase.product_MRP),
            'purchase_rate': purchase_rate,
            'sale_rate': 0.0,
            'cgst': cgst,
            'sgst': sgst,
            'gst_amount': gst_amount,
            'purchase_cost': purchase_cost,
            'sales_value': 0.0,
            'profit': -purchase_cost,
            'profit_percentage': 0.0
        })
        
        total_purchase_cost += purchase_cost
        total_profit -= purchase_cost
        total_gst += gst_amount
    
    # Process Supplier Challans (only non-invoiced)
    for challan in supplier_challan_list:
        quantity = float(challan.product_quantity)
        purchase_rate = float(challan.product_purchase_rate)
        cgst = float(challan.cgst)
        sgst = float(challan.sgst)
        
        purchase_cost = purchase_rate * quantity
        gst_amount = (cgst + sgst) * quantity
        
        financial_data.append({
            'type': 'Supplier Challan',
            'date': challan.challan_entry_date,
            'invoice_no': challan.product_challan_no,
            'customer': challan.product_suppliername.supplier_name,
            'product_name': challan.product_name,
            'company': challan.product_company,
            'batch_no': challan.product_batch_no,
            'quantity': quantity,
            'mrp': float(challan.product_mrp),
            'purchase_rate': purchase_rate,
            'sale_rate': 0.0,
            'cgst': cgst,
            'sgst': sgst,
            'gst_amount': gst_amount,
            'purchase_cost': purchase_cost,
            'sales_value': 0.0,
            'profit': -purchase_cost,
            'profit_percentage': 0.0
        })
        
        total_purchase_cost += purchase_cost
        total_profit -= purchase_cost
        total_gst += gst_amount
    
    # Stock valuation removed for performance
    stock_valuation = 0.0
    
    # Summary statistics
    summary = {
        'total_sales_value': total_sales_value,
        'total_purchase_cost': total_purchase_cost,
        'total_gst': total_gst,
        'total_profit': total_profit,
        'profit_percentage': (total_profit / total_sales_value * 100) if total_sales_value > 0 else 0,
        'total_transactions': len(financial_data),
        'stock_valuation': stock_valuation
    }
    
    # Pagination - 20 records per page
    from django.core.paginator import Paginator
    paginator = Paginator(financial_data, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'financial_data': page_obj,
        'summary': summary,
        'start_date': start_date,
        'end_date': end_date,
        'selected_product': product_id,
        'product_search': product_search
    }
    
    return render(request, 'reports/financial_report.html', context)


@login_required
def export_financial_pdf(request):
    """Export financial report as PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from io import BytesIO
    
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    product_id = request.GET.get('product_id')
    
    # Get data
    sales_query = SalesMaster.objects.select_related('productid', 'sales_invoice_no', 'customerid')
    supplier_challan_query = SupplierChallanMaster.objects.select_related('product_id', 'product_suppliername', 'product_challan_id')
    customer_challan_query = CustomerChallanMaster.objects.select_related('product_id', 'customer_name', 'customer_challan_id')
    
    if start_date:
        try:
            sales_query = sales_query.filter(sale_entry_date__date__gte=start_date)
            supplier_challan_query = supplier_challan_query.filter(challan_entry_date__date__gte=start_date)
            customer_challan_query = customer_challan_query.filter(sales_entry_date__date__gte=start_date)
        except:
            pass
    if end_date:
        try:
            sales_query = sales_query.filter(sale_entry_date__date__lte=end_date)
            supplier_challan_query = supplier_challan_query.filter(challan_entry_date__date__lte=end_date)
            customer_challan_query = customer_challan_query.filter(sales_entry_date__date__lte=end_date)
        except:
            pass
    if product_id and product_id.strip():
        try:
            pid = int(product_id)
            sales_query = sales_query.filter(productid_id=pid)
            supplier_challan_query = supplier_challan_query.filter(product_id=pid)
            customer_challan_query = customer_challan_query.filter(product_id=pid)
        except:
            pass
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#1a237e'), alignment=1)
    elements.append(Paragraph('Financial Report - Profit Analysis', title_style))
    elements.append(Spacer(1, 12))
    
    # Date range
    if start_date or end_date:
        date_text = f"Period: {start_date or 'Start'} to {end_date or 'End'}"
        elements.append(Paragraph(date_text, styles['Normal']))
        elements.append(Spacer(1, 12))
    
    # Table data
    data = [['Date', 'Type', 'Invoice', 'Party', 'Product', 'Batch', 'Qty', 'P.Rate', 'S.Rate', 'GST', 'Profit']]
    
    total_profit = 0.0
    total_sales = 0.0
    all_txns = []
    
    # Collect all transactions
    for sale in sales_query:
        purchase = PurchaseMaster.objects.filter(productid=sale.productid, product_batch_no=sale.product_batch_no).first()
        purchase_rate = float(purchase.product_purchase_rate) if purchase else 0.0
        quantity = float(sale.sale_quantity)
        sale_rate = float(sale.sale_rate)
        gst_amount = (float(sale.sale_cgst) + float(sale.sale_sgst)) * quantity
        purchase_cost = purchase_rate * quantity
        sales_value = sale_rate * quantity
        profit = (sales_value - purchase_cost) + gst_amount
        all_txns.append({
            'date': sale.sale_entry_date, 'type': 'Sale', 'invoice': sale.sales_invoice_no.sales_invoice_no,
            'party': sale.customerid.customer_name, 'product': sale.product_name[:15], 'batch': sale.product_batch_no,
            'qty': quantity, 'p_rate': purchase_rate, 's_rate': sale_rate, 'gst': gst_amount, 'profit': profit, 'sales': sales_value
        })
    
    for challan in customer_challan_query:
        purchase = PurchaseMaster.objects.filter(productid=challan.product_id, product_batch_no=challan.product_batch_no).first()
        purchase_rate = float(purchase.product_purchase_rate) if purchase else 0.0
        quantity = float(challan.sale_quantity)
        sale_rate = float(challan.sale_rate)
        gst_amount = (float(challan.sale_cgst) + float(challan.sale_sgst)) * quantity
        purchase_cost = purchase_rate * quantity
        sales_value = sale_rate * quantity
        profit = (sales_value - purchase_cost) + gst_amount
        all_txns.append({
            'date': challan.sales_entry_date, 'type': 'C.Challan', 'invoice': challan.customer_challan_no,
            'party': challan.customer_name.customer_name, 'product': challan.product_name[:15], 'batch': challan.product_batch_no,
            'qty': quantity, 'p_rate': purchase_rate, 's_rate': sale_rate, 'gst': gst_amount, 'profit': profit, 'sales': sales_value
        })
    
    for challan in supplier_challan_query:
        quantity = float(challan.product_quantity)
        purchase_rate = float(challan.product_purchase_rate)
        gst_amount = (float(challan.cgst) + float(challan.sgst)) * quantity
        purchase_cost = purchase_rate * quantity
        profit = -purchase_cost
        all_txns.append({
            'date': challan.challan_entry_date, 'type': 'S.Challan', 'invoice': challan.product_challan_no,
            'party': challan.product_suppliername.supplier_name, 'product': challan.product_name[:15], 'batch': challan.product_batch_no,
            'qty': quantity, 'p_rate': purchase_rate, 's_rate': 0, 'gst': gst_amount, 'profit': profit, 'sales': 0
        })
    
    # Sort and limit
    all_txns.sort(key=lambda x: x['date'], reverse=True)
    for txn in all_txns[:100]:
        data.append([
            txn['date'].strftime('%d-%m-%Y'), txn['type'], txn['invoice'], txn['party'], txn['product'], txn['batch'],
            f"{txn['qty']:.0f}", f"₹{txn['p_rate']:.2f}", f"₹{txn['s_rate']:.2f}", f"₹{txn['gst']:.2f}", f"₹{txn['profit']:.2f}"
        ])
        total_profit += txn['profit']
        total_sales += txn['sales']
    
    # Add summary row
    data.append(['', '', '', '', '', '', '', '', '', 'TOTAL:', f'₹{total_profit:.2f}'])
    
    # Create table
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8eaf6')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="financial_report_{datetime.now().strftime("%Y%m%d")}.pdf"'
    return response


@login_required
def export_financial_excel(request):
    """Export financial report as Excel"""
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    product_id = request.GET.get('product_id')
    
    # Get data
    sales_query = SalesMaster.objects.select_related('productid', 'sales_invoice_no', 'customerid')
    supplier_challan_query = SupplierChallanMaster.objects.select_related('product_id', 'product_suppliername', 'product_challan_id')
    customer_challan_query = CustomerChallanMaster.objects.select_related('product_id', 'customer_name', 'customer_challan_id')
    
    if start_date:
        try:
            sales_query = sales_query.filter(sale_entry_date__date__gte=start_date)
            supplier_challan_query = supplier_challan_query.filter(challan_entry_date__date__gte=start_date)
            customer_challan_query = customer_challan_query.filter(sales_entry_date__date__gte=start_date)
        except:
            pass
    if end_date:
        try:
            sales_query = sales_query.filter(sale_entry_date__date__lte=end_date)
            supplier_challan_query = supplier_challan_query.filter(challan_entry_date__date__lte=end_date)
            customer_challan_query = customer_challan_query.filter(sales_entry_date__date__lte=end_date)
        except:
            pass
    if product_id and product_id.strip():
        try:
            pid = int(product_id)
            sales_query = sales_query.filter(productid_id=pid)
            supplier_challan_query = supplier_challan_query.filter(product_id=pid)
            customer_challan_query = customer_challan_query.filter(product_id=pid)
        except:
            pass
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Report"
    
    # Header styling
    header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Date', 'Type', 'Invoice No', 'Party', 'Product', 'Company', 'Batch', 'Qty', 
               'MRP', 'Purchase Rate', 'Sale Rate', 'CGST', 'SGST', 'GST Amount', 
               'Purchase Cost', 'Sales Value', 'Profit', 'Profit %']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    row_num = 2
    total_sales_value = 0.0
    total_purchase_cost = 0.0
    total_profit = 0.0
    total_gst = 0.0
    
    # Process Sales
    for sale in sales_query:
        purchase = PurchaseMaster.objects.filter(productid=sale.productid, product_batch_no=sale.product_batch_no).first()
        purchase_rate = float(purchase.product_purchase_rate) if purchase else 0.0
        quantity = float(sale.sale_quantity)
        sale_rate = float(sale.sale_rate)
        gst_amount = (float(sale.sale_cgst) + float(sale.sale_sgst)) * quantity
        purchase_cost = purchase_rate * quantity
        sales_value = sale_rate * quantity
        profit = (sales_value - purchase_cost) + gst_amount
        profit_pct = (profit / sales_value * 100) if sales_value > 0 else 0
        
        row_data = [
            sale.sale_entry_date.strftime('%d-%m-%Y'), 'Sale', sale.sales_invoice_no.sales_invoice_no,
            sale.customerid.customer_name, sale.product_name, sale.product_company, sale.product_batch_no,
            quantity, float(sale.product_MRP), purchase_rate, sale_rate, float(sale.sale_cgst), float(sale.sale_sgst),
            gst_amount, purchase_cost, sales_value, profit, profit_pct
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if col_num >= 8:
                cell.alignment = Alignment(horizontal='right')
        
        total_sales_value += sales_value
        total_purchase_cost += purchase_cost
        total_profit += profit
        total_gst += gst_amount
        row_num += 1
    
    # Process Customer Challans
    for challan in customer_challan_query:
        purchase = PurchaseMaster.objects.filter(productid=challan.product_id, product_batch_no=challan.product_batch_no).first()
        purchase_rate = float(purchase.product_purchase_rate) if purchase else 0.0
        quantity = float(challan.sale_quantity)
        sale_rate = float(challan.sale_rate)
        gst_amount = (float(challan.sale_cgst) + float(challan.sale_sgst)) * quantity
        purchase_cost = purchase_rate * quantity
        sales_value = sale_rate * quantity
        profit = (sales_value - purchase_cost) + gst_amount
        profit_pct = (profit / sales_value * 100) if sales_value > 0 else 0
        
        row_data = [
            challan.sales_entry_date.strftime('%d-%m-%Y'), 'Customer Challan', challan.customer_challan_no,
            challan.customer_name.customer_name, challan.product_name, challan.product_company, challan.product_batch_no,
            quantity, float(challan.product_mrp), purchase_rate, sale_rate, float(challan.sale_cgst), float(challan.sale_sgst),
            gst_amount, purchase_cost, sales_value, profit, profit_pct
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if col_num >= 8:
                cell.alignment = Alignment(horizontal='right')
        
        total_sales_value += sales_value
        total_purchase_cost += purchase_cost
        total_profit += profit
        total_gst += gst_amount
        row_num += 1
    
    # Process Supplier Challans
    for challan in supplier_challan_query:
        quantity = float(challan.product_quantity)
        purchase_rate = float(challan.product_purchase_rate)
        gst_amount = (float(challan.cgst) + float(challan.sgst)) * quantity
        purchase_cost = purchase_rate * quantity
        profit = -purchase_cost
        
        row_data = [
            challan.challan_entry_date.strftime('%d-%m-%Y'), 'Supplier Challan', challan.product_challan_no,
            challan.product_suppliername.supplier_name, challan.product_name, challan.product_company, challan.product_batch_no,
            quantity, float(challan.product_mrp), purchase_rate, 0, float(challan.cgst), float(challan.sgst),
            gst_amount, purchase_cost, 0, profit, 0
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if col_num >= 8:
                cell.alignment = Alignment(horizontal='right')
        
        total_purchase_cost += purchase_cost
        total_profit -= purchase_cost
        total_gst += gst_amount
        row_num += 1
    
    # Summary row
    summary_fill = PatternFill(start_color="e8eaf6", end_color="e8eaf6", fill_type="solid")
    summary_font = Font(bold=True, size=11)
    
    summary_row = row_num + 1
    ws.cell(row=summary_row, column=1).value = "TOTAL"
    ws.cell(row=summary_row, column=1).font = summary_font
    ws.cell(row=summary_row, column=14).value = round(total_gst, 2)
    ws.cell(row=summary_row, column=15).value = round(total_purchase_cost, 2)
    ws.cell(row=summary_row, column=16).value = round(total_sales_value, 2)
    ws.cell(row=summary_row, column=17).value = round(total_profit, 2)
    ws.cell(row=summary_row, column=18).value = round((total_profit / total_sales_value * 100), 2) if total_sales_value > 0 else 0
    
    for col in range(1, 19):
        cell = ws.cell(row=summary_row, column=col)
        cell.fill = summary_fill
        cell.font = summary_font
        cell.border = border
    
    # Adjust column widths
    for col in range(1, 19):
        ws.column_dimensions[get_column_letter(col)].width = 14
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="financial_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    
    return response
